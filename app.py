"""
NSE Market Dashboard — plain-language stock analysis for everyday investors.
Run:  python app.py   →   open http://127.0.0.1:8050
"""
import sys, io, json, base64, subprocess, tempfile, os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import dash
from dash import dcc, html, Input, Output, State, ctx, dash_table
import dash_bootstrap_components as dbc

from config import (NSE_TICKERS, DATA_CLEANED, DATA_FEATURES, DATA_RAW,
                    DEFAULT_INVESTMENT, DEFAULT_CONFIDENCE, MONTE_CARLO_HORIZON)

# ── Company metadata ─────────────────────────────────────────────────────────
_PALETTE = [
    "#38bdf8","#a78bfa","#f472b6","#fb923c","#34d399","#f59e0b",
    "#60a5fa","#e879f9","#4ade80","#fbbf24","#c084fc","#818cf8",
    "#f87171","#2dd4bf","#facc15","#94a3b8","#06b6d4","#8b5cf6",
    "#ec4899","#84cc16",
]
_SECTOR_ICONS = {
    "Agricultural":                     "🌾",
    "Automobiles and Accessories":      "🚗",
    "Banking":                          "🏦",
    "Commercial and Services":          "🛒",
    "Construction and Allied":          "🏗️",
    "Energy and Petroleum":             "⚡",
    "Insurance":                        "🛡️",
    "Investment":                       "💼",
    "Manufacturing and Allied":         "🏭",
    "Media":                            "📰",
    "Real Estate Investment Trust":     "🏠",
    "Telecommunication and Technology": "📱",
    "Transport and Storage":            "✈️",
}
_RAW_COMPANIES = [
    # (ticker_short, name, NSE sector)
    ("ABSA", "ABSA Bank Kenya",              "Banking"),
    ("ALP",  "ALP Industrial REIT",          "Real Estate Investment Trust"),
    ("AMAC", "Africa Mega Agricorp",         "Agricultural"),
    ("BAT",  "BAT Kenya",                    "Manufacturing and Allied"),
    ("BKG",  "BK Group",                     "Banking"),
    ("BOC",  "BOC Kenya",                    "Energy and Petroleum"),
    ("BRIT", "Britam Holdings",              "Insurance"),
    ("CARB", "Carbacid Investments",         "Construction and Allied"),
    ("CGEN", "Centum Generation",            "Investment"),
    ("CIC",  "CIC Insurance Group",          "Insurance"),
    ("COOP", "Co-operative Bank",            "Banking"),
    ("CRWN", "Crown Paints Kenya",           "Construction and Allied"),
    ("CTUM", "Cavendish Management",         "Commercial and Services"),
    ("DTK",  "Diamond Trust Bank",           "Banking"),
    ("EABL", "East African Breweries",       "Manufacturing and Allied"),
    ("EGAD", "East African Portland Cement", "Investment"),
    ("EQTY", "Equity Group Holdings",        "Banking"),
    ("EVRD", "Eveready East Africa",         "Commercial and Services"),
    ("FMLY", "Family Bank",                  "Banking"),
    ("FTGH", "Fahari I-REIT",               "Real Estate Investment Trust"),
    ("GLD",  "Gold Coin Kenya",              "Commercial and Services"),
    ("HAFR", "Home Afrika",                  "Commercial and Services"),
    ("HFCK", "HF Group",                     "Banking"),
    ("IMH",  "I&M Holdings",                 "Banking"),
    ("JUB",  "Jubilee Holdings",             "Insurance"),
    ("KAPC", "KAPS Medical International",   "Commercial and Services"),
    ("KCB",  "KCB Group",                    "Banking"),
    ("KEGN", "KenGen",                       "Energy and Petroleum"),
    ("KNRE", "Kenya Reinsurance",            "Insurance"),
    ("KPC",  "Kenya Power (Ord)",            "Energy and Petroleum"),
    ("KPLC", "Kenya Power (Pref)",           "Energy and Petroleum"),
    ("KQ",   "Kenya Airways",               "Transport and Storage"),
    ("KUKZ", "Kakuzi",                       "Agricultural"),
    ("KURV", "Kurwitu Ventures",             "Commercial and Services"),
    ("LBTY", "Liberty Kenya Holdings",       "Insurance"),
    ("LIMT", "Limuru Tea",                   "Agricultural"),
    ("LKL",  "Longhorn Publishers",          "Commercial and Services"),
    ("NBV",  "Nairobi Business Ventures",    "Banking"),
    ("NCBA", "NCBA Group",                   "Banking"),
    ("NMG",  "Nation Media Group",           "Media"),
    ("NSE",  "Nairobi Securities Exchange",  "Investment"),
    ("OCH",  "Olympia Capital Holdings",     "Commercial and Services"),
    ("PORT", "East African Portland Cement", "Construction and Allied"),
    ("SASN", "Sasini",                       "Agricultural"),
    ("SBIC", "SBM Bank Kenya",               "Banking"),
    ("SCAN", "Scangroup",                    "Commercial and Services"),
    ("SCBK", "Standard Chartered Bank Kenya","Banking"),
    ("SCOM", "Safaricom",                    "Telecommunication and Technology"),
    ("SGL",  "Standard Group",               "Media"),
    ("SKL",  "Stanbic Kenya",               "Banking"),
    ("SLAM", "Sanlam Kenya",                 "Insurance"),
    ("SMER", "Sameer Africa",               "Automobiles and Accessories"),
    ("SMWF", "Stanlib Fahari REIT",          "Real Estate Investment Trust"),
    ("TOTL", "TotalEnergies EP Kenya",       "Energy and Petroleum"),
    ("TPSE", "TransCentury",                 "Transport and Storage"),
    ("TRFC", "TransAfrica",                  "Transport and Storage"),
    ("UCHM", "Unga Group Chemicals",         "Manufacturing and Allied"),
    ("UMME", "Umme",                         "Investment"),
    ("UNGA", "Unga Group",                   "Manufacturing and Allied"),
    ("WTK",  "Williamson Tea Kenya",         "Agricultural"),
    ("XPRS", "Express Kenya",               "Commercial and Services"),
]
# Build COMPANIES — keep the original 5 companies' colors/icons intact
_ORIGINAL_COLORS = {
    "SCOM": "#38bdf8", "EQTY": "#a78bfa",
    "KCB":  "#f472b6", "EABL": "#fb923c", "COOP": "#34d399",
}
COMPANIES = {}
_palette_idx = 0
for _short, _name, _sector in _RAW_COMPANIES:
    _ticker = f"{_short}.NR"
    if _short in _ORIGINAL_COLORS:
        _color = _ORIGINAL_COLORS[_short]
    else:
        _color = _PALETTE[_palette_idx % len(_PALETTE)]
        _palette_idx += 1
    COMPANIES[_ticker] = dict(
        name=_name, short=_short, sector=_sector,
        color=_color, icon=_SECTOR_ICONS.get(_sector, "📈"),
    )
# Override icons for original 5
COMPANIES["SCOM.NR"]["icon"] = "📱"
COMPANIES["EQTY.NR"]["icon"] = "🏦"
COMPANIES["KCB.NR"]["icon"]  = "🏦"
COMPANIES["EABL.NR"]["icon"] = "🍺"
COMPANIES["COOP.NR"]["icon"] = "🏦"

# Auto-derive SECTORS from COMPANIES
_all_sectors = {}
for _tk, _meta in COMPANIES.items():
    _all_sectors.setdefault(_meta["sector"], []).append(_tk)
SECTORS = {"All Companies": list(COMPANIES.keys()), **_all_sectors}

PERIODS = {"1 Month":21,"3 Months":63,"6 Months":126,"1 Year":252,"3 Years":756,"All Time":9999,"Custom Range":0}

# ── Design tokens ────────────────────────────────────────────────────────────
C = dict(bg="#07090f", panel="#0d1117", card="#161b22", border="#21262d",
         accent="#38bdf8", buy="#22c55e", sell="#ef4444", hold="#f59e0b",
         text="#e6edf3", muted="#8b949e", header="#010409")

ADVICE = {
    "BUY":  dict(label="Good time to buy",   color=C["buy"],  bg="#052e16", emoji="↑"),
    "SELL": dict(label="Consider selling",   color=C["sell"], bg="#2d0a0a", emoji="↓"),
    "HOLD": dict(label="Hold what you have", color=C["hold"], bg="#292101", emoji="→"),
}
CHART_BASE = dict(template="plotly_dark", paper_bgcolor=C["card"], plot_bgcolor=C["card"],
                  font=dict(color=C["text"], size=11),
                  margin=dict(l=10, r=10, t=40, b=10),
                  legend=dict(bgcolor="rgba(0,0,0,0)"))

# ── Data helpers ──────────────────────────────────────────────────────────────
_ARCHIVE_DIR = Path(r"C:\Users\moeng\Downloads\archive")


def _normalise_archive_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise column names: strip whitespace and lowercase, then rename to canonical."""
    df.columns = [c.strip() for c in df.columns]
    lower_map = {c: c.lower() for c in df.columns}
    df = df.rename(columns=lower_map)
    canonical = {"date": "Date", "code": "Code", "day price": "Day Price", "volume": "Volume"}
    return df.rename(columns=canonical)


def _build_archive_master() -> pd.DataFrame:
    """Load all NSE archive CSVs (+ daily patch files) into a single dataframe.

    Handles both old (DATE/CODE uppercase, 2000-2021) and new (Date/Code, 2022+) formats.
    """
    frames = []

    def _load_chunk(path: Path) -> None:
        try:
            raw = pd.read_csv(path, dtype=str)
            df  = _normalise_archive_cols(raw)
            if "Code" not in df.columns or "Day Price" not in df.columns:
                return
            df["Code"] = df["Code"].str.strip()
            df["_dt"] = pd.to_datetime(
                df["Date"].str.strip(), dayfirst=False, format="mixed", errors="coerce"
            )
            df = df.dropna(subset=["_dt"])
            df["Close"] = pd.to_numeric(
                df["Day Price"].str.replace(",", "", regex=False).str.strip(),
                errors="coerce",
            )
            df["Volume"] = pd.to_numeric(
                df["Volume"].str.replace(",", "", regex=False).str.strip(),
                errors="coerce",
            ) if "Volume" in df.columns else np.nan
            df = df[df["Close"] > 0].dropna(subset=["Close"])
            frames.append(df[["_dt", "Code", "Close", "Volume"]])
        except Exception:
            pass

    for path in sorted(_ARCHIVE_DIR.glob("NSE_data_all_stocks_????.csv")):
        _load_chunk(path)
    for patch in sorted(_ARCHIVE_DIR.glob("NSE_patch_*.csv")):
        _load_chunk(patch)

    if not frames:
        return pd.DataFrame(columns=["_dt", "Code", "Close", "Volume"])
    master = pd.concat(frames, ignore_index=True)
    master = master.sort_values(["Code", "_dt"])
    master = master.drop_duplicates(subset=["_dt", "Code"], keep="last")
    return master


print("Loading NSE archive data…", flush=True)
_ARCHIVE_MASTER: pd.DataFrame = _build_archive_master()
_ARCHIVE_CODES: set = set(_ARCHIVE_MASTER["Code"].unique()) if not _ARCHIVE_MASTER.empty else set()
print(f"  Archive ready: {len(_ARCHIVE_MASTER):,} rows across {len(_ARCHIVE_CODES)} companies.", flush=True)

_archive_df_cache: dict = {}


def _load_company_archive(code: str):
    """Return time-series df with Close column for a single NSE code from master archive."""
    if code in _archive_df_cache:
        return _archive_df_cache[code]
    if _ARCHIVE_MASTER.empty or code not in _ARCHIVE_CODES:
        _archive_df_cache[code] = None
        return None
    try:
        rows = _ARCHIVE_MASTER[_ARCHIVE_MASTER["Code"] == code].copy()
        df = pd.DataFrame({"Close": rows["Close"].values, "Volume": rows["Volume"].values},
                          index=rows["_dt"])
        df = df[~df.index.duplicated(keep="last")].sort_index()
        df["Close"] = _clean_price_series(df["Close"])
        result = df if len(df) >= 5 else None
        _archive_df_cache[code] = result
        return result
    except Exception:
        _archive_df_cache[code] = None
        return None


def _repair_decimal_errors(s: pd.Series) -> pd.Series:
    """Fix decimal-point-dropped prices using a trailing-year reference window.

    Why trailing window instead of global median:
    - Some NSE stocks have genuinely changed price scale over 20 years (e.g. a company
      at 25 KES in 2006 that declined to 0.68 KES today). Global median would flag
      the 2006 legitimate prices as corrupt.
    - Trailing window (preceding ~1 year) gives a clean local reference: for 2026
      corrupt data, it compares against 2025 correct prices; for 2006 data, it
      compares against other 2006 prices.

    Algorithm (iterated until stable):
    1. For each price, compute the log10 median of the preceding 252 rows.
       Fall back to global log10 median when fewer than 20 preceding rows exist.
    2. Flag values that are more than 20x ABOVE the trailing reference.
    3. Among factors 10 / 100 / 1000, pick the one whose result is CLOSEST to
       the reference median (avoids over-dividing 185 → 0.185 when 1.85 is correct).
    4. Repeat until convergence so that runs of consecutive corrupt rows are fully
       resolved in successive passes.
    """
    arr = s.values.astype(float).copy()
    n = len(arr)

    valid = arr[(arr > 0) & ~np.isnan(arr)]
    if len(valid) < 10:
        return s

    LOG_THRESH  = np.log10(20)   # 20x from reference → suspect
    TRAIL       = 252            # ≈ 1 trading year look-back
    MIN_CONTEXT = 20             # minimum trailing rows for reliable median

    # Global fallback: used only for the first MIN_CONTEXT rows
    global_log_med = float(np.median(np.log10(valid)))

    # Track which indices the main loop modifies so the fine-tune pass can
    # restrict itself to those rows only — never touching values the main loop
    # left alone (those are within the 20x threshold and considered legitimate).
    main_modified: set[int] = set()

    for _ in range(30):
        changed = False
        for i in range(n):
            val = arr[i]
            if np.isnan(val) or val <= 0:
                continue

            lo = max(0, i - TRAIL)
            ctx = arr[lo:i]
            ctx = ctx[(ctx > 0) & ~np.isnan(ctx)]
            log_med = float(np.median(np.log10(ctx))) if len(ctx) >= MIN_CONTEXT else global_log_med

            log_val = np.log10(val)
            if log_val - log_med <= LOG_THRESH:
                continue

            best_candidate, best_dist = None, float("inf")
            for factor in (1000.0, 100.0, 10.0):
                candidate = val / factor
                dist = abs(np.log10(candidate) - log_med)
                if dist < LOG_THRESH and dist < best_dist:
                    best_dist, best_candidate = dist, candidate

            if best_candidate is not None:
                arr[i] = best_candidate
                changed = True
                main_modified.add(i)

        if not changed:
            break

    # Fine-tune pass: only refines rows the main loop already changed.
    # Bidirectional context prevents confusing a genuine price-level step
    # change with corruption. Iterating lets later-cleaned neighbours
    # unblock rows that were skipped in earlier passes.
    FTUNE_LOCAL  = 5
    FTUNE_THRESH = np.log10(8)
    for _ in range(10):
        ftune_changed = False
        for i in sorted(main_modified):
            val = arr[i]
            if np.isnan(val) or val <= 0:
                continue
            lo = max(0, i - FTUNE_LOCAL)
            hi = min(n, i + FTUNE_LOCAL + 1)
            ctx_b = arr[lo:i]
            ctx_a = arr[i + 1:hi]
            ctx_b = ctx_b[(ctx_b > 0) & ~np.isnan(ctx_b)]
            ctx_a = ctx_a[(ctx_a > 0) & ~np.isnan(ctx_a)]
            if len(ctx_b) < 3 or len(ctx_a) < 3:
                continue
            log_val     = np.log10(val)
            log_med_b   = float(np.median(np.log10(ctx_b)))
            log_med_a   = float(np.median(np.log10(ctx_a)))
            if log_val - log_med_b <= FTUNE_THRESH or log_val - log_med_a <= FTUNE_THRESH:
                continue
            log_med_local = (log_med_b + log_med_a) / 2
            best_candidate, best_dist = None, float("inf")
            for factor in (1000.0, 100.0, 10.0):
                candidate = val / factor
                dist = abs(np.log10(candidate) - log_med_local)
                if dist < FTUNE_THRESH and dist < best_dist:
                    best_dist, best_candidate = dist, candidate
            if best_candidate is not None:
                arr[i] = best_candidate
                ftune_changed = True
        if not ftune_changed:
            break

    return pd.Series(arr, index=s.index)


def _clean_price_series(s: pd.Series) -> pd.Series:
    """Repair decimal-dropped outliers, then cap remaining spikes with log-space IQR."""
    s = _repair_decimal_errors(s)
    if len(s) < 10:
        return s
    log_s = np.log(s.clip(lower=1e-6))
    q10, q90 = log_s.quantile(0.10), log_s.quantile(0.90)
    spread = q90 - q10
    if spread <= 0:
        return s
    lower = np.exp(q10 - 3 * spread)
    upper = np.exp(q90 + 3 * spread)
    return s.clip(lower=max(lower, 0.0), upper=upper)


_load_df_cache: dict = {}


def load_df(ticker: str):
    if ticker in _load_df_cache:
        return _load_df_cache[ticker]
    code = ticker.split(".")[0].upper()
    p = DATA_CLEANED / f"{ticker.replace('.','_')}_cleaned.csv"
    if p.exists():
        df = pd.read_csv(p, index_col="Date", parse_dates=True)
        # Replace OHLC with archive-repaired prices (decimal corruption fixed there).
        arc = _load_company_archive(code)
        if arc is not None and not arc.empty and "Close" in arc.columns:
            raw_close = df["Close"].copy()
            arc_close = arc["Close"].reindex(df.index)
            valid = arc_close.notna()
            if valid.any():
                df.loc[valid, "Close"] = arc_close[valid]
                ratio = arc_close.div(raw_close.replace(0, np.nan)).fillna(1.0).clip(1e-4, 1e4)
                for col in ["Open", "High", "Low"]:
                    if col in df.columns:
                        df.loc[valid, col] = (df[col].mul(ratio)).loc[valid]
                # Pass 2: for dates not in archive that are still clearly corrupt
                # (>20x from the archive-corrected median), interpolate from neighbors.
                if not valid.all():
                    close = df["Close"].astype(float)
                    log_close = np.log(close.clip(lower=1e-9))
                    log_med = float(np.median(log_close[valid]))
                    corrupt_orphan = ~valid & (np.abs(log_close - log_med) > np.log(20))
                    if corrupt_orphan.any():
                        for col in ["Close", "Open", "High", "Low"]:
                            if col in df.columns:
                                s = df[col].astype(float).copy()
                                s[corrupt_orphan] = np.nan
                                s = s.interpolate(method="time").ffill().bfill()
                                df[col] = s
        _load_df_cache[ticker] = df
        return df
    result = _load_company_archive(code)
    _load_df_cache[ticker] = result
    return result


def _compute_ta_signal(df) -> dict:
    """Compute basic technical analysis signal when no ML signal is available."""
    close = df["Close"].dropna()
    if len(close) < 10:
        return {"risk_adjusted_signal": "—"}
    last = float(close.iloc[-1])
    n = len(close)
    ma20 = float(close.tail(min(20, n)).mean())
    ma50 = float(close.tail(min(50, n)).mean())
    delta = close.diff().dropna()
    gain = delta.clip(lower=0).tail(14).mean()
    loss = (-delta.clip(upper=0)).tail(14).mean()
    rsi = round(100 - (100 / (1 + gain / loss)), 1) if loss and loss != 0 else 50.0
    mom20 = round((close.iloc[-1] - close.iloc[max(-20,-n)]) / close.iloc[max(-20,-n)] * 100, 2)
    if last > ma20 and ma20 >= ma50 and rsi < 70:
        signal = "BUY"
    elif last < ma20 and ma20 <= ma50 and rsi > 30:
        signal = "SELL"
    else:
        signal = "HOLD"
    return {
        "risk_adjusted_signal": signal,
        "signal_source": "technical",
        "rsi_14": rsi,
        "ma20": round(ma20, 2),
        "ma50": round(ma50, 2),
        "momentum_20d_pct": mom20,
        "var_95_pct": round(abs(mom20) / 3, 2),
    }


def load_sig(ticker: str):
    p = DATA_FEATURES / f"{ticker.replace('.','_')}_signal.json"
    if p.exists():
        with open(p) as f:
            sig = json.load(f)
            sig.setdefault("signal_source", "ml")
            return sig
    df = load_df(ticker)
    return _compute_ta_signal(df) if df is not None else None

def pct_change_over(df, days):
    n = min(days, len(df)-1)
    if n < 1: return 0.0
    return (df["Close"].iloc[-1] - df["Close"].iloc[-1-n]) / df["Close"].iloc[-1-n] * 100

def risk_label(var_pct):
    v = abs(var_pct or 0)
    if v < 5:   return "Low",    C["buy"]
    if v < 10:  return "Medium", C["hold"]
    return "High", C["sell"]

# ── Chart builders ────────────────────────────────────────────────────────────
def chart_price_simple(df, ticker, days=252):
    meta = COMPANIES.get(ticker, {})
    name = meta.get("name", ticker)
    color = meta.get("color", C["accent"])
    sub = df.tail(min(days, len(df)))
    fig = go.Figure()
    # Fill area under price line
    fig.add_trace(go.Scatter(
        x=sub.index, y=sub["Close"], name="Price",
        line=dict(color=color, width=2),
        fill="tozeroy", fillcolor=color.replace(")", ",0.08)").replace("rgb","rgba") if "rgb" in color else f"rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.08)",
    ))
    # Simple 50-day average line
    if len(sub) > 50:
        avg = sub["Close"].rolling(50).mean()
        fig.add_trace(go.Scatter(x=sub.index, y=avg, name="50-day average",
                                 line=dict(color="#ffffff", width=1, dash="dot"), opacity=0.5))
    fig.update_layout(**CHART_BASE, height=320,
                      title=dict(text=f"{name} — Price History (KES)", font=dict(color=color)),
                      xaxis_title="", yaxis_title="Price (KES)")
    fig.update_xaxes(gridcolor=C["border"], zeroline=False)
    fig.update_yaxes(gridcolor=C["border"], zeroline=False)
    return fig


def chart_returns_bar(df, name, color):
    periods = {"1M":21,"3M":63,"6M":126,"1Y":252,"3Y":756}
    labels, values, colors = [], [], []
    for lbl, d in periods.items():
        if len(df) > d:
            v = pct_change_over(df, d)
            labels.append(lbl); values.append(round(v,2))
            colors.append(C["buy"] if v >= 0 else C["sell"])
    fig = go.Figure(go.Bar(x=labels, y=values, marker_color=colors,
                           text=[f"{v:+.1f}%" for v in values], textposition="outside"))
    fig.update_layout(**CHART_BASE, height=260,
                      title=dict(text=f"{name} — Returns by Period", font=dict(color=color)),
                      yaxis_title="Change (%)", showlegend=False)
    fig.update_xaxes(gridcolor=C["border"]); fig.update_yaxes(gridcolor=C["border"])
    return fig


def _slice(series, days=252, start=None, end=None):
    """Slice a pandas Series by date range or trailing day count."""
    if start and end:
        return series[start:end]
    return series.tail(min(days, len(series)))


def chart_comparison_indexed(tickers, days=252, start=None, end=None):
    fig = go.Figure()
    for t in tickers:
        df = load_df(t)
        if df is None or len(df) < 2: continue
        meta = COMPANIES.get(t, {})
        sub = _slice(df["Close"], days, start, end)
        if sub.empty: continue
        base = sub.iloc[0]
        if base == 0: continue
        indexed = (sub / base * 100).round(2)
        fig.add_trace(go.Scatter(
            x=indexed.index, y=indexed.values,
            name=meta.get("name", t),
            line=dict(color=meta.get("color", C["accent"]), width=2),
        ))
    fig.add_hline(y=100, line_dash="dash", line_color=C["muted"],
                  annotation_text="Starting point", annotation_position="left")
    fig.update_layout(**CHART_BASE, height=400,
                      title=dict(text="How KES 100 invested in each company would have grown",
                                 font=dict(color=C["accent"])),
                      yaxis_title="Value of KES 100 invested",
                      hovermode="x unified")
    fig.update_xaxes(gridcolor=C["border"]); fig.update_yaxes(gridcolor=C["border"])
    return fig


def chart_performance_ranked(tickers=None, days=252, start=None, end=None):
    source = tickers if tickers is not None else list(COMPANIES.keys())
    rows = []
    for t in source:
        meta = COMPANIES.get(t)
        if meta is None: continue
        df = load_df(t)
        if df is None: continue
        if start and end:
            sub = df["Close"][start:end]
            if len(sub) < 2: continue
            chg = (sub.iloc[-1] - sub.iloc[0]) / sub.iloc[0] * 100 if sub.iloc[0] else 0.0
        else:
            chg = pct_change_over(df, days)
        rows.append((meta["name"], chg, meta["color"]))
    rows.sort(key=lambda x: x[1])
    names, vals, colors = zip(*rows) if rows else ([], [], [])
    bar_colors = [C["buy"] if v >= 0 else C["sell"] for v in vals]
    fig = go.Figure(go.Bar(
        y=list(names), x=list(vals),
        orientation="h", marker_color=bar_colors,
        text=[f"{v:+.1f}%" for v in vals], textposition="auto",
    ))
    fig.update_layout(**{**CHART_BASE, "margin": dict(l=160, r=10, t=40, b=10)},
                      height=280,
                      title=dict(text="Which company performed best?", font=dict(color=C["accent"])),
                      xaxis_title="Price Change (%)", showlegend=False)
    fig.update_xaxes(gridcolor=C["border"]); fig.update_yaxes(gridcolor=C["border"])
    return fig


def chart_monthly_heatmap(ticker):
    df = load_df(ticker)
    meta = COMPANIES.get(ticker, {})
    name = meta.get("name", ticker) if ticker else "—"
    if df is None:
        fig = go.Figure()
        fig.update_layout(**CHART_BASE, height=300,
                          title=dict(text=f"No cleaned data for {name}",
                                     font=dict(color=C["muted"])))
        fig.add_annotation(
            text="Import data for this company via the Import & Analyse tab,<br>"
                 "or pick a different company above.",
            xref="paper", yref="paper", x=0.5, y=0.5,
            showarrow=False, font=dict(color=C["muted"], size=13),
        )
        return fig
    monthly = df["Close"].resample("ME").last().pct_change().dropna() * 100
    monthly.index = pd.to_datetime(monthly.index)
    years  = sorted(monthly.index.year.unique())
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    z = []
    for yr in years:
        row = []
        for m in range(1,13):
            vals = monthly[(monthly.index.year == yr) & (monthly.index.month == m)]
            row.append(round(float(vals.iloc[0]),2) if len(vals) > 0 else None)
        z.append(row)
    fig = go.Figure(go.Heatmap(
        z=z, x=months, y=[str(yr) for yr in years],
        colorscale=[[0,"#7f1d1d"],[0.5,"#1e293b"],[1,"#14532d"]],
        zmid=0,
        text=[[f"{v:+.1f}%" if v is not None else "" for v in row] for row in z],
        texttemplate="%{text}",
        colorbar=dict(title="Monthly\nReturn %"),
    ))
    fig.update_layout(**{**CHART_BASE, "margin": dict(l=50, r=80, t=40, b=10)},
                      height=380,
                      title=dict(text=f"{meta.get('name',ticker)} — Monthly Returns Calendar",
                                 font=dict(color=meta.get("color", C["accent"]))))
    return fig


def chart_all_prices(tickers, days=252, start=None, end=None):
    # Pre-filter to tickers that actually have loaded data
    available = [(t, load_df(t)) for t in tickers]
    available = [(t, df) for t, df in available if df is not None]
    if not available:
        return go.Figure()
    n = len(available)
    spacing = min(0.04, 0.8 / max(n - 1, 1))
    fig = make_subplots(
        rows=n, cols=1, shared_xaxes=True,
        subplot_titles=[COMPANIES.get(t, {}).get("name", t) for t, _ in available],
        vertical_spacing=spacing,
    )
    for i, (t, df) in enumerate(available, 1):
        meta  = COMPANIES.get(t, {})
        sub   = _slice(df["Close"], days, start, end)
        if sub.empty: continue
        color = meta.get("color", C["accent"])
        fig.add_trace(go.Scatter(
            x=sub.index, y=sub.values, name=meta.get("name", t),
            line=dict(color=color, width=1.5),
            fill="tozeroy",
            fillcolor=(lambda c: f"rgba({int(c[1:3],16)},{int(c[3:5],16)},{int(c[5:7],16)},0.09)")(color),
            showlegend=False,
        ), row=i, col=1)
    fig.update_layout(**CHART_BASE, height=max(120 * n + 40, 300),
                      title=dict(text="Price History — All Companies with data (KES)",
                                 font=dict(color=C["accent"])))
    fig.update_xaxes(gridcolor=C["border"], zeroline=False)
    fig.update_yaxes(gridcolor=C["border"], zeroline=False)
    return fig

# ── UI helpers ────────────────────────────────────────────────────────────────
def advice_badge(sig, size="sm"):
    a = ADVICE.get(sig, dict(label=sig or "—", color=C["muted"], bg=C["card"], emoji=""))
    fs = {"sm":"0.68rem","md":"0.9rem","lg":"1.3rem"}.get(size,"0.68rem")
    px = {"sm":"3px 9px","md":"6px 14px","lg":"10px 22px"}.get(size,"3px 9px")
    return html.Span(f"{a['emoji']} {a['label']}", style=dict(
        background=a["bg"], color=a["color"], border=f"1px solid {a['color']}",
        fontWeight=700, fontSize=fs, padding=px, borderRadius="20px",
        letterSpacing="0.02em", whiteSpace="nowrap",
    ))

def stat_pill(label, value, color=None):
    return html.Div([
        html.Div(label, style=dict(fontSize="0.62rem", color=C["muted"],
                                   textTransform="uppercase", letterSpacing="0.07em")),
        html.Div(value, style=dict(fontSize="0.95rem", fontWeight=700,
                                   color=color or C["text"])),
    ], style=dict(background=C["card"], border=f"1px solid {C['border']}",
                  borderRadius="8px", padding="8px 14px", textAlign="center"))

# ── App ───────────────────────────────────────────────────────────────────────
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.CYBORG],
                suppress_callback_exceptions=True)
app.title = "NSE Market Dashboard"

# Inject date-picker CSS directly into the HTML head — works regardless of
# which react-dates class names the installed Dash version uses.
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
        /* ── Date picker: target the raw input directly via wrapper ID ── */
        #explorer-dates-wrap input {
            color: #e6edf3 !important;
            background-color: #161b22 !important;
            font-size: 0.82rem !important;
            font-weight: 500 !important;
            border: none !important;
            outline: none !important;
        }
        #explorer-dates-wrap input::placeholder { color: #8b949e !important; }

        /* Also cover analytics single-date pickers */
        #analytics-custom-start input,
        #analytics-custom-end input {
            color: #e6edf3 !important;
            background-color: #161b22 !important;
            font-size: 0.82rem !important;
            font-weight: 500 !important;
            border: none !important;
        }

        /* react-dates class names (belt-and-suspenders) */
        .DateInput_input {
            color: #e6edf3 !important;
            background-color: #161b22 !important;
            font-size: 0.82rem !important;
            font-weight: 500 !important;
        }
        .DateInput_input::placeholder { color: #8b949e !important; }
        .DateRangePickerInput, .SingleDatePickerInput {
            background-color: #161b22 !important;
            border: 1px solid #21262d !important;
            border-radius: 6px !important;
        }
        .DateInput { background: transparent !important; }

        /* Calendar popup */
        .DayPicker, .CalendarMonthGrid, .CalendarMonth {
            background: #161b22 !important;
            color: #e6edf3 !important;
        }
        .CalendarMonth_caption, .CalendarMonth_caption strong { color: #e6edf3 !important; }
        .DayPicker_weekHeader_li small { color: #8b949e !important; }
        .CalendarDay__default {
            background: #161b22 !important;
            border: 1px solid #21262d !important;
            color: #e6edf3 !important;
        }
        .CalendarDay__default:hover { background: #21262d !important; }
        .CalendarDay__selected, .CalendarDay__selected:active, .CalendarDay__selected:hover {
            background: #38bdf8 !important;
            border-color: #38bdf8 !important;
            color: #010409 !important;
            font-weight: 700 !important;
        }
        .CalendarDay__selected_span {
            background: #0c4a6e !important;
            border-color: #0369a1 !important;
            color: #e6edf3 !important;
        }
        .DayPickerNavigation_button__default {
            background: #161b22 !important;
            border: 1px solid #21262d !important;
        }
        .DayPickerNavigation_svg__horizontal { fill: #8b949e !important; }
        .DateRangePickerInput_arrow_svg,
        .DateRangePickerInput_arrow { color: #8b949e !important; fill: #8b949e !important; }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

# Tab style helpers
TAB_STYLE = dict(background=C["panel"], color=C["muted"], border="none",
                 padding="10px 20px", fontSize="0.85rem")
TAB_SELECTED = dict(background=C["card"], color=C["accent"],
                    border=f"1px solid {C['border']}", borderBottom="none",
                    padding="10px 20px", fontSize="0.85rem", fontWeight=700)

app.layout = html.Div([
    dcc.Store(id="selected-ticker", data="SCOM.NR"),
    dcc.Store(id="analysis-store", data={}),
    dcc.Store(id="analytics-state", data={"days":252,"sector":"All Companies","heatmap":"SCOM.NR"}),
    dcc.Store(id="overview-view",   data="table"),
    dcc.Store(id="overview-sector", data="All Companies"),
    dcc.Store(id="pipeline-ticker",    data=""),
    dcc.Store(id="pipeline-csv-path",  data=""),
    dcc.Store(id="explorer-data",      data={}),
    dcc.Store(id="explorer-chart-type", data="line"),
    dcc.Download(id="explorer-download-csv"),
    dcc.Download(id="explorer-download-excel"),
    # Interval for polling full-pipeline status (disabled by default)
    dcc.Interval(id="pipeline-poll", interval=3000, n_intervals=0, disabled=True),

    # ── Header ───────────────────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.Span("📈", style=dict(fontSize="1.3rem", marginRight="8px")),
            html.Span("NSE Market Dashboard",
                      style=dict(fontWeight=800, fontSize="1.1rem", color=C["text"])),
            html.Span("Nairobi Securities Exchange · KES",
                      style=dict(fontSize="0.75rem", color=C["muted"], marginLeft="12px")),
        ], style=dict(display="flex", alignItems="center")),
        html.Div([
            dcc.Input(id="search-input", placeholder="🔍  Search company…",
                      debounce=True, style=dict(
                          background=C["card"], border=f"1px solid {C['border']}",
                          color=C["text"], borderRadius="20px",
                          padding="7px 16px", width="230px", fontSize="0.85rem")),
        ], style=dict(display="flex", gap="10px", alignItems="center")),
    ], style=dict(background=C["header"], borderBottom=f"1px solid {C['border']}",
                  padding="12px 24px", display="flex",
                  justifyContent="space-between", alignItems="center")),

    # ── Main tabs ─────────────────────────────────────────────────────────────
    dcc.Tabs(id="main-tabs", value="overview", children=[
        dcc.Tab(label="🏠  Overview",        value="overview",   style=TAB_STYLE, selected_style=TAB_SELECTED),
        dcc.Tab(label="🔍  Company",          value="company",    style=TAB_STYLE, selected_style=TAB_SELECTED),
        dcc.Tab(label="📊  Analytics",        value="analytics",  style=TAB_STYLE, selected_style=TAB_SELECTED),
        dcc.Tab(label="📥  Import & Analyse", value="import_tab", style=TAB_STYLE, selected_style=TAB_SELECTED),
        dcc.Tab(label="📅  Data Explorer",   value="explorer",   style=TAB_STYLE, selected_style=TAB_SELECTED),
    ], style=dict(background=C["panel"], borderBottom=f"1px solid {C['border']}")),

    # Permanent overview controls bar — always in DOM, hidden on other tabs
    html.Div([
        html.Div([
            html.Span("Sector: ", style=dict(color=C["muted"], fontSize="0.8rem", marginRight="6px")),
            dcc.Dropdown(
                id="ov-sector-dropdown",
                options=[{"label": k, "value": k} for k in SECTORS],
                value="All Companies",
                clearable=False,
                style=dict(background=C["card"], color=C["text"], fontSize="0.82rem",
                           border=f"1px solid {C['border']}", borderRadius="8px",
                           width="220px", minWidth="180px"),
            ),
            # Hidden buttons kept for callback compatibility (no_clicks, never shown)
            html.Button(id="ov-s-all",       n_clicks=0, style=dict(display="none")),
            html.Button(id="ov-s-banking",   n_clicks=0, style=dict(display="none")),
            html.Button(id="ov-s-telecom",   n_clicks=0, style=dict(display="none")),
            html.Button(id="ov-s-beverages", n_clicks=0, style=dict(display="none")),
        ], style=dict(display="flex", alignItems="center", flexWrap="wrap", gap="8px")),
        html.Div([
            html.Button("☰  Table", id="ov-v-table", n_clicks=0,
                        style=dict(border=f"1px solid {C['border']}", borderRadius="20px",
                                   padding="5px 16px", cursor="pointer", fontSize="0.8rem",
                                   fontWeight=600, background=C["accent"], color=C["header"])),
            html.Button("⬛  Board", id="ov-v-board", n_clicks=0,
                        style=dict(border=f"1px solid {C['border']}", borderRadius="20px",
                                   padding="5px 16px", cursor="pointer", fontSize="0.8rem",
                                   fontWeight=600, background=C["card"], color=C["muted"])),
        ], style=dict(display="flex", gap="6px")),
    ], id="overview-controls-bar",
       style=dict(display="none", justifyContent="space-between", alignItems="center",
                  padding="10px 24px", borderBottom=f"1px solid {C['border']}")),

    dcc.Loading(html.Div(id="tab-content"), type="dot", color=C["accent"]),

], style=dict(background=C["bg"], color=C["text"], minHeight="100vh",
              fontFamily="'Inter','Segoe UI',sans-serif", margin=0, padding=0))

# ── Tab content router ────────────────────────────────────────────────────────
@app.callback(Output("tab-content","children"),
              Input("main-tabs","value"),
              Input("selected-ticker","data"),
              State("overview-sector","data"),
              State("overview-view","data"),
              State("analysis-store","data"),
              State("analytics-state","data"))
def render_tab(tab, ticker, ov_sector, ov_view, store, astate):
    astate    = astate or {"days":252,"sector":"All Companies","heatmap":"SCOM.NR"}
    ov_sector = ov_sector or "All Companies"
    ov_view   = ov_view   or "table"
    if tab == "overview":
        return html.Div(id="overview-body",
                        children=_build_overview_content(ov_sector, ov_view))
    if tab == "company":    return build_company(ticker, store or {})
    if tab == "analytics":  return build_analytics(
        astate["days"], astate["sector"], astate.get("heatmap","SCOM.NR"),
        astate.get("custom_start"), astate.get("custom_end"))
    if tab == "import_tab": return build_import_tab()
    if tab == "explorer":   return build_explorer_tab()
    return html.Div()


# ── Overview controls bar: show/hide based on active tab ─────────────────────
@app.callback(
    Output("overview-controls-bar","style"),
    Input("main-tabs","value"),
)
def toggle_controls_bar(tab):
    vis = dict(display="flex", justifyContent="space-between", alignItems="center",
               padding="10px 24px", borderBottom=f"1px solid {C['border']}")
    return vis if tab == "overview" else dict(display="none")


# ── Overview sector filter (fixed IDs — no pattern matching) ─────────────────
@app.callback(
    Output("overview-sector","data"),
    Input("ov-sector-dropdown","value"),
    prevent_initial_call=True,
)
def set_ov_sector(sector):
    return sector or "All Companies"


# ── Overview view toggle (fixed IDs) ─────────────────────────────────────────
@app.callback(
    Output("overview-view","data"),
    Input("ov-v-table","n_clicks"),
    Input("ov-v-board","n_clicks"),
    prevent_initial_call=True,
)
def set_ov_view(n_table, n_board):
    return "board" if ctx.triggered_id == "ov-v-board" else "table"


# ── Update overview body when sector or view changes ─────────────────────────
@app.callback(
    Output("overview-body","children"),
    Input("overview-sector","data"),
    Input("overview-view","data"),
    prevent_initial_call=True,
)
def update_overview_body(sector, view):
    return _build_overview_content(sector or "All Companies", view or "table")





# ── Highlight active view button ─────────────────────────────────────────────
@app.callback(
    Output("ov-v-table","style"),
    Output("ov-v-board","style"),
    Input("overview-view","data"),
)
def update_view_btn_styles(view):
    view = view or "table"
    return [
        dict(**_btn_base, background=C["accent"] if view=="table" else C["card"],
             color=C["header"] if view=="table" else C["muted"]),
        dict(**_btn_base, background=C["accent"] if view=="board" else C["card"],
             color=C["header"] if view=="board" else C["muted"]),
    ]

@app.callback(
    Output("analytics-custom-range-section", "style"),
    Input("analytics-period", "value"),
    prevent_initial_call=True,
)
def toggle_custom_range(period):
    visible = dict(display="flex", gap="14px", alignItems="flex-start", flexWrap="wrap",
                   marginTop="10px", padding="10px 24px",
                   borderTop=f"1px solid {C['border']}")
    hidden  = dict(display="none")
    return visible if period == 0 else hidden


@app.callback(
    Output("analytics-state","data"),
    Input("analytics-period","value"),
    Input("analytics-sector","value"),
    Input("heatmap-ticker","value"),
    Input("analytics-apply-range","n_clicks"),
    State("analytics-custom-start","date"),
    State("analytics-custom-end","date"),
    State("analytics-state","data"),
    prevent_initial_call=True,
)
def update_analytics_state(period, sector, heatmap, n_apply, custom_start, custom_end, state):
    state = state or {"days":252,"sector":"All Companies","heatmap":"SCOM.NR","custom_start":None,"custom_end":None}
    triggered = ctx.triggered_id
    if triggered == "analytics-apply-range" and custom_start and custom_end:
        return {**state, "days":0, "sector":sector or state["sector"],
                "heatmap":heatmap or state["heatmap"],
                "custom_start":custom_start, "custom_end":custom_end}
    if triggered != "analytics-apply-range":
        return {**state,
                "days":period if period is not None else state["days"],
                "sector":sector or state["sector"],
                "heatmap":heatmap or state["heatmap"]}
    return state

# ── TAB 1: OVERVIEW — Today's Signal Board ────────────────────────────────────
def _build_overview_content(sector="All Companies", view="table"):
    # Gather data for signal table — filtered by sector
    filtered_tickers = SECTORS.get(sector, list(COMPANIES.keys()))
    rows = []
    buy_count = sell_count = hold_count = 0
    for t in filtered_tickers:
        meta = COMPANIES.get(t, {})
        df   = load_df(t)
        sig  = load_sig(t)
        if df is None:
            continue
        last  = float(df["Close"].iloc[-1])
        prev  = float(df["Close"].iloc[-2]) if len(df) > 1 else last
        chg   = (last - prev) / prev * 100 if prev else 0.0
        chg1y = pct_change_over(df, 252)
        ra    = sig.get("risk_adjusted_signal", "—") if sig else "—"
        rlabel, _ = risk_label(sig.get("var_95_pct", 0) if sig else 0)

        if ra == "BUY":  buy_count  += 1
        elif ra == "SELL": sell_count += 1
        elif ra == "HOLD": hold_count += 1

        rows.append({
            "_ticker": t,
            "Company":        meta.get("name", t),
            "Sector":         meta.get("sector", "—"),
            "Price (KES)":    f"{last:,.2f}",
            "Today's Change": f"{chg:+.2f}%",
            "1-Year Change":  f"{chg1y:+.1f}%",
            "AI Advice":      ra,
            "Risk Level":     rlabel,
        })

    # Counter badges
    counter_style = lambda bg, border: dict(
        background=bg, border=f"1px solid {border}",
        borderRadius="12px", padding="16px 28px", textAlign="center",
        minWidth="120px",
    )
    counters = html.Div([
        html.Div([
            html.Div(str(buy_count),  style=dict(fontSize="2rem", fontWeight=800, color=C["buy"])),
            html.Div("BUY signals",   style=dict(fontSize="0.72rem", color=C["muted"], marginTop="2px")),
        ], style=counter_style("#052e1680", C["buy"])),
        html.Div([
            html.Div(str(hold_count), style=dict(fontSize="2rem", fontWeight=800, color=C["hold"])),
            html.Div("HOLD signals",  style=dict(fontSize="0.72rem", color=C["muted"], marginTop="2px")),
        ], style=counter_style("#29210180", C["hold"])),
        html.Div([
            html.Div(str(sell_count), style=dict(fontSize="2rem", fontWeight=800, color=C["sell"])),
            html.Div("SELL signals",  style=dict(fontSize="0.72rem", color=C["muted"], marginTop="2px")),
        ], style=counter_style("#2d0a0a80", C["sell"])),
    ], style=dict(display="flex", gap="16px", flexWrap="wrap",
                  padding="16px 24px", borderBottom=f"1px solid {C['border']}"))

    # ── Board view (Kanban columns: BUY | HOLD | SELL) ────────────────────────
    def company_mini_card(r):
        adv = r["AI Advice"]
        meta = COMPANIES.get(r["_ticker"], {})
        adv_color, adv_bg = {
            "BUY":  (C["buy"],  "#052e16"),
            "SELL": (C["sell"], "#2d0a0a"),
            "HOLD": (C["hold"], "#292101"),
        }.get(adv, (C["muted"], C["card"]))
        chg_val = r["Today's Change"]
        chg_c = C["buy"] if "+" in chg_val and chg_val != "+0.00%" else C["sell"] if "-" in chg_val else C["muted"]
        return html.Div([
            html.Div([
                html.Span(meta.get("icon",""), style=dict(fontSize="1.1rem")),
                html.Div([
                    html.Div(r["Company"], style=dict(fontWeight=700, fontSize="0.85rem", color=C["text"])),
                    html.Span(r["Sector"], style=dict(fontSize="0.63rem", color=meta.get("color",C["accent"]),
                                                       background=meta.get("color","#fff")+"22",
                                                       padding="1px 7px", borderRadius="10px")),
                ]),
            ], style=dict(display="flex", alignItems="center", gap="8px", marginBottom="10px")),
            html.Div([
                html.Span(f"KES {r['Price (KES)']}", style=dict(fontWeight=800, fontSize="0.95rem", color=C["text"])),
                html.Span(f" {chg_val}", style=dict(color=chg_c, fontSize="0.78rem", fontWeight=600)),
            ], style=dict(marginBottom="4px")),
            html.Div(f"1Y: {r['1-Year Change']}  ·  Risk: {r['Risk Level']}",
                     style=dict(fontSize="0.7rem", color=C["muted"])),
        ], id={"type":"overview-row","index":r["_ticker"]}, n_clicks=0,
           style=dict(background=C["card"], border=f"1px solid {adv_bg}",
                      borderRadius="10px", padding="12px", cursor="pointer",
                      marginBottom="10px", transition="border 0.15s"))

    def kanban_col(title, signal, col_color, col_bg, col_rows):
        cards = [company_mini_card(r) for r in col_rows if r["AI Advice"] == signal]
        empty = html.Div("No companies here", style=dict(color=C["muted"], fontSize="0.8rem",
                                                          fontStyle="italic", textAlign="center",
                                                          padding="20px")) if not cards else None
        return html.Div([
            html.Div([
                html.Span(f"{'↑' if signal=='BUY' else '↓' if signal=='SELL' else '→'}  {signal}",
                          style=dict(fontWeight=800, fontSize="0.88rem", color=col_color)),
                html.Span(str(len(cards)),
                          style=dict(background=col_bg, color=col_color, border=f"1px solid {col_color}",
                                     borderRadius="12px", padding="1px 9px", fontSize="0.75rem",
                                     fontWeight=700, marginLeft="8px")),
            ], style=dict(display="flex", alignItems="center",
                          marginBottom="14px", paddingBottom="10px",
                          borderBottom=f"1px solid {col_color}40")),
            html.Div(cards if cards else [empty]),
        ], style=dict(flex="1", background=C["panel"], border=f"1px solid {col_color}30",
                      borderRadius="12px", padding="16px", minWidth="220px"))

    board_view = html.Div([
        kanban_col("Good time to buy",   "BUY",  C["buy"],  "#052e16", rows),
        kanban_col("Hold what you have", "HOLD", C["hold"], "#292101", rows),
        kanban_col("Consider selling",   "SELL", C["sell"], "#2d0a0a", rows),
    ], id="overview-board-view",
       style=dict(display="flex" if view=="board" else "none",
                  gap="16px", padding="16px 24px 24px", flexWrap="wrap"))

    # ── Signal table rows (manual HTML — dash_table doesn't support clickable cell content easily) ──
    def signal_color(sig_val):
        return {
            "BUY":  (C["buy"],  "#052e16"),
            "SELL": (C["sell"], "#2d0a0a"),
            "HOLD": (C["hold"], "#292101"),
        }.get(sig_val, (C["muted"], C["card"]))

    def chg_color(chg_str):
        try:
            v = float(chg_str.replace("%","").replace("+",""))
            return C["buy"] if v >= 0 else C["sell"]
        except Exception:
            return C["muted"]

    th_style = dict(
        padding="10px 14px", fontSize="0.72rem", color=C["muted"],
        fontWeight=700, textTransform="uppercase", letterSpacing="0.07em",
        borderBottom=f"2px solid {C['border']}", textAlign="left",
        background=C["panel"],
    )
    td_style = dict(
        padding="10px 14px", fontSize="0.83rem", color=C["text"],
        borderBottom=f"1px solid {C['border']}",
    )

    header_row = html.Tr([
        html.Th("Company",        style=th_style),
        html.Th("Sector",         style=th_style),
        html.Th("Price (KES)",    style={**th_style, "textAlign":"right"}),
        html.Th("Today's Change", style={**th_style, "textAlign":"right"}),
        html.Th("1-Year Change",  style={**th_style, "textAlign":"right"}),
        html.Th("AI Advice",      style={**th_style, "textAlign":"center"}),
        html.Th("Risk Level",     style={**th_style, "textAlign":"center"}),
    ])

    table_rows = []
    for r in rows:
        t = r["_ticker"]
        adv_color, adv_bg = signal_color(r["AI Advice"])
        risk_col = {
            "Low":    C["buy"],
            "Medium": C["hold"],
            "High":   C["sell"],
        }.get(r["Risk Level"], C["muted"])

        table_rows.append(html.Tr([
            # Clickable company name
            html.Td(
                html.Span(r["Company"], id={"type":"overview-row","index":t},
                          n_clicks=0, style=dict(
                              color=C["accent"], cursor="pointer", fontWeight=600,
                              textDecoration="underline", fontSize="0.85rem",
                          )),
                style=td_style,
            ),
            html.Td(r["Sector"],        style=td_style),
            html.Td(r["Price (KES)"],   style={**td_style, "textAlign":"right", "fontVariantNumeric":"tabular-nums"}),
            html.Td(r["Today's Change"],style={**td_style, "textAlign":"right",
                                              "color": chg_color(r["Today's Change"]),
                                              "fontWeight":600}),
            html.Td(r["1-Year Change"], style={**td_style, "textAlign":"right",
                                              "color": chg_color(r["1-Year Change"]),
                                              "fontWeight":600}),
            html.Td(
                html.Span(f"{ADVICE.get(r['AI Advice'],{}).get('emoji','')} {r['AI Advice']}",
                          style=dict(background=adv_bg, color=adv_color,
                                     border=f"1px solid {adv_color}",
                                     padding="3px 10px", borderRadius="20px",
                                     fontSize="0.72rem", fontWeight=700,
                                     whiteSpace="nowrap")),
                style={**td_style, "textAlign":"center"},
            ),
            html.Td(
                html.Span(r["Risk Level"], style=dict(color=risk_col, fontWeight=700)),
                style={**td_style, "textAlign":"center"},
            ),
        ], style=dict(transition="background 0.12s"), className="overview-row"))

    signal_table = html.Div([
        html.Table(
            [html.Thead(header_row), html.Tbody(table_rows)],
            style=dict(width="100%", borderCollapse="collapse"),
        )
    ], style=dict(overflowX="auto",
                  border=f"1px solid {C['border']}", borderRadius="10px",
                  background=C["card"]))

    hint = html.Div(
        "Click any company name to see full details on the Company tab.",
        style=dict(fontSize="0.78rem", color=C["muted"], fontStyle="italic",
                   marginTop="8px", textAlign="right"),
    )

    return html.Div([
        # intro banner
        html.Div([
            html.Div("Today's Signal Board", style=dict(
                fontSize="1.3rem", fontWeight=800, color=C["text"], marginBottom="4px")),
            html.Div("What should I do today across all my investments? Our AI gives you a simple signal for every company.",
                     style=dict(fontSize="0.85rem", color=C["muted"])),
        ], style=dict(padding="20px 24px 12px", borderBottom=f"1px solid {C['border']}")),

        # BUY / HOLD / SELL counters
        counters,

        # Table view (hidden when board is active)
        html.Div([
            signal_table,
            hint,
        ], id="overview-table-view",
           style=dict(padding="16px 24px 24px",
                      display="none" if view=="board" else "block")),

        # Board view (hidden when table is active)
        board_view,

        # Legend
        html.Div([
            html.Div("How to read the signals:", style=dict(
                fontWeight=700, color=C["text"], marginBottom="8px", fontSize="0.85rem")),
            html.Div([
                html.Div([advice_badge("BUY","sm"),
                          html.Span(" — Our AI thinks the stock price will rise. Could be a good time to invest.",
                                    style=dict(color=C["muted"], fontSize="0.8rem"))],
                         style=dict(display="flex", alignItems="center", gap="8px", marginBottom="6px")),
                html.Div([advice_badge("SELL","sm"),
                          html.Span(" — Our AI thinks the price may fall. Consider reducing your position.",
                                    style=dict(color=C["muted"], fontSize="0.8rem"))],
                         style=dict(display="flex", alignItems="center", gap="8px", marginBottom="6px")),
                html.Div([advice_badge("HOLD","sm"),
                          html.Span(" — The AI is not confident enough to recommend buying or selling right now.",
                                    style=dict(color=C["muted"], fontSize="0.8rem"))],
                         style=dict(display="flex", alignItems="center", gap="8px")),
            ]),
        ], style=dict(margin="0 24px 24px", background=C["card"],
                      border=f"1px solid {C['border']}", borderRadius="10px",
                      padding="16px")),
    ], style=dict(overflowY="auto", maxHeight="calc(100vh - 105px)"))


# ── Callbacks for overview clickable rows ────────────────────────────────────
@app.callback(
    Output("selected-ticker","data"),
    Output("main-tabs","value"),
    Input({"type":"overview-row","index":dash.ALL},"n_clicks"),
    Input({"type":"sidebar-ticker","index":dash.ALL},"n_clicks"),
    Input("search-input","value"),
    State("selected-ticker","data"),
    State("main-tabs","value"),
    prevent_initial_call=True,
)
def select_ticker_and_navigate(overview_clicks, sidebar_clicks, search, current_ticker, current_tab):
    t = ctx.triggered_id
    if isinstance(t, dict):
        new_ticker = t["index"]
        if t.get("type") == "overview-row":
            return new_ticker, "company"
        else:
            # sidebar click — stay on company tab
            return new_ticker, "company"
    if t == "search-input" and search:
        s = search.strip().upper()
        tk = s if s.endswith(".NR") else s + ".NR"
        return tk, "company"
    return current_ticker, current_tab


# ── TAB 2: COMPANY — Deep Dive ────────────────────────────────────────────────
def build_company(ticker, store):
    # Sidebar
    sidebar_items = []
    for t, meta in COMPANIES.items():
        df  = load_df(t)
        sig = load_sig(t)
        last = float(df["Close"].iloc[-1]) if df is not None else 0
        prev = float(df["Close"].iloc[-2]) if df is not None and len(df)>1 else last
        chg  = (last-prev)/prev*100 if prev else 0
        ra   = sig.get("risk_adjusted_signal","—") if sig else "—"
        a    = ADVICE.get(ra, dict(color=C["muted"], bg=C["card"]))
        is_active = t == ticker
        sidebar_items.append(html.Div([
            html.Div(meta["icon"], style=dict(fontSize="1.1rem", width="28px", textAlign="center")),
            html.Div([
                html.Div(meta["name"], style=dict(fontSize="0.82rem", fontWeight=600,
                                                  color=C["accent"] if is_active else C["text"])),
                html.Div(f"KES {last:,.2f}  {chg:+.1f}%",
                         style=dict(fontSize="0.7rem",
                                    color=C["buy"] if chg>=0 else C["sell"])),
            ], style=dict(flex=1)),
            html.Span(f"{ADVICE.get(ra,{}).get('emoji','')} {ra}",
                      style=dict(fontSize="0.65rem", fontWeight=700, color=a["color"],
                                 background=a["bg"], border=f"1px solid {a['color']}",
                                 padding="2px 7px", borderRadius="10px")),
        ], id={"type":"sidebar-ticker","index":t}, n_clicks=0,
           style=dict(display="flex", alignItems="center", gap="8px",
                      padding="10px 14px", cursor="pointer",
                      background="#1c2333" if is_active else "transparent",
                      borderBottom=f"1px solid {C['border']}")))

    sidebar = html.Div([
        html.Div("Companies", style=dict(fontSize="0.65rem", fontWeight=700, color=C["muted"],
                                         letterSpacing="0.1em", padding="10px 14px",
                                         borderBottom=f"1px solid {C['border']}")),
        html.Div(sidebar_items),
    ], style=dict(width="260px", minWidth="260px", background=C["panel"],
                  borderRight=f"1px solid {C['border']}", overflowY="auto"))

    # Main content
    meta = COMPANIES.get(ticker, {})
    df   = load_df(ticker)
    sig  = store.get(ticker) or load_sig(ticker)

    if df is None:
        main = html.Div("No data available for this company.",
                        style=dict(color=C["muted"], padding="40px"))
    else:
        last  = float(df["Close"].iloc[-1])
        prev  = float(df["Close"].iloc[-2]) if len(df)>1 else last
        chg   = (last-prev)/prev*100
        chg_c = C["buy"] if chg>=0 else C["sell"]

        ra_sig = sig.get("risk_adjusted_signal","—") if sig else "—"
        signal = sig.get("signal","—") if sig else "—"
        pred   = sig.get("predicted_price_KES",0) if sig else 0
        pred_c = sig.get("predicted_change_pct",0) if sig else 0
        var_p  = sig.get("var_95_pct",0) if sig else 0
        acc    = sig["metrics"]["directional_accuracy"] if sig and sig.get("metrics") else None
        rlabel, rcolor = risk_label(var_p)
        worst_kes = abs(var_p/100) * DEFAULT_INVESTMENT

        main_children = [
            # Company title
            html.Div([
                html.Span(meta.get("icon",""), style=dict(fontSize="2rem", marginRight="12px")),
                html.Div([
                    html.Div(meta.get("name", ticker),
                             style=dict(fontSize="1.4rem", fontWeight=800, color=C["text"])),
                    html.Span(meta.get("sector",""), style=dict(
                        fontSize="0.7rem", color=meta.get("color",C["accent"]),
                        background=meta.get("color","#fff")+"22",
                        padding="2px 8px", borderRadius="10px")),
                ]),
            ], style=dict(display="flex", alignItems="center",
                          padding="20px 24px 12px")),

            # Price strip
            html.Div([
                html.Div([
                    html.Span(f"KES {last:,.2f}",
                              style=dict(fontSize="2rem", fontWeight=800, color=C["text"])),
                    html.Span(f"  {chg:+.2f}% today",
                              style=dict(color=chg_c, fontSize="1rem", fontWeight=600)),
                ]),
                html.Div(f"52-week range:  KES {df['Close'].tail(252).min():,.2f} – KES {df['Close'].tail(252).max():,.2f}",
                         style=dict(fontSize="0.78rem", color=C["muted"], marginTop="4px")),
            ], style=dict(padding="0 24px 16px",
                          borderBottom=f"1px solid {C['border']}")),

            # Recommendation section
            html.Div([
                html.Div("🤖  What our AI recommends", style=dict(
                    fontWeight=700, fontSize="1rem", color=C["text"], marginBottom="16px")),
                html.Div([
                    # Big advice badge
                    html.Div([
                        html.Div("Our Advice", style=dict(fontSize="0.65rem", color=C["muted"],
                                                          marginBottom="6px", letterSpacing="0.08em")),
                        advice_badge(ra_sig, "lg"),
                        html.Div(ADVICE.get(ra_sig,{}).get("label",""), style=dict(
                            fontSize="0.75rem", color=C["muted"], marginTop="6px")),
                    ], style=dict(textAlign="center", padding="16px",
                                  background=C["card"], borderRadius="10px",
                                  border=f"1px solid {ADVICE.get(ra_sig,{}).get('color',C['border'])}")),

                    # AI reasoning in plain English
                    html.Div([
                        html.Div("Why?", style=dict(fontWeight=700, color=C["text"],
                                                     fontSize="0.85rem", marginBottom="8px")),
                        html.Div([
                            html.Span("📌 ", style=dict(fontSize="0.9rem")),
                            html.Span(
                                f"Our AI predicts the price will move from "
                                f"KES {last:,.2f} to about KES {pred:,.2f} "
                                f"({'rising' if pred_c>0 else 'falling'} by {abs(pred_c):.1f}%).",
                                style=dict(fontSize="0.83rem", color=C["muted"])),
                        ], style=dict(marginBottom="8px")),
                        html.Div([
                            html.Span("⚠️ " if rlabel=="High" else "✅ ", style=dict(fontSize="0.9rem")),
                            html.Span(
                                f"Risk level: {rlabel}. If you invest KES 100,000 in {meta.get('name','this company')}, "
                                f"on a bad day you could lose up to KES {worst_kes:,.0f}.",
                                style=dict(fontSize="0.83rem", color=C["muted"])),
                        ], style=dict(marginBottom="8px")),
                        html.Div([
                            html.Span("🎯 ", style=dict(fontSize="0.9rem")),
                            html.Span(
                                f"Our AI has correctly predicted price direction "
                                f"{acc:.0f} out of every 100 times." if acc else
                                "Run an analysis to see AI accuracy.",
                                style=dict(fontSize="0.83rem", color=C["muted"])),
                        ]),
                    ], style=dict(background=C["card"], borderRadius="10px",
                                  padding="14px", border=f"1px solid {C['border']}")),

                    # Key numbers
                    html.Div([
                        stat_pill("Today's Price",    f"KES {last:,.2f}"),
                        stat_pill("AI Price Target",  f"KES {pred:,.2f}", C["accent"]),
                        stat_pill("Predicted Change", f"{pred_c:+.1f}%",
                                  C["buy"] if pred_c>0 else C["sell"]),
                        stat_pill("Risk Level",       rlabel, rcolor),
                        stat_pill("AI Accuracy",      f"{acc:.0f}%" if acc else "—", C["accent"]),
                        stat_pill("1-Year Change",    f"{pct_change_over(df,252):+.1f}%",
                                  C["buy"] if pct_change_over(df,252)>=0 else C["sell"]),
                    ], style=dict(display="flex", gap="8px", flexWrap="wrap")),

                ], style=dict(display="flex", flexDirection="column", gap="14px")),

                html.Div([
                    html.Button("🔄 Update Analysis", id="quick-btn", style=dict(
                        background=C["accent"], color=C["header"], fontWeight=700,
                        border="none", borderRadius="20px", padding="7px 20px",
                        cursor="pointer", fontSize="0.83rem", marginTop="14px")),
                    dcc.Loading(html.Span(id="quick-status"), type="circle", color=C["accent"]),
                ], style=dict(display="flex", alignItems="center", gap="12px")),

            ], style=dict(padding="20px 24px",
                          borderBottom=f"1px solid {C['border']}")),

            # Price chart with period selector
            html.Div([
                html.Div([
                    html.Span("📈 Price History", style=dict(fontWeight=700, fontSize="0.9rem")),
                    dcc.Dropdown(
                        id="price-period",
                        options=[{"label":k,"value":v} for k,v in PERIODS.items()],
                        value=252,
                        clearable=False,
                        style=dict(background=C["card"], color=C["text"],
                                   border=f"1px solid {C['border']}", width="140px",
                                   fontSize="0.8rem"),
                    ),
                ], style=dict(display="flex", justifyContent="space-between",
                              alignItems="center", marginBottom="8px")),
                dcc.Graph(id="price-chart-company",
                          figure=chart_price_simple(df, ticker),
                          config=dict(displayModeBar=False, scrollZoom=True)),
            ], style=dict(padding="16px 24px", borderBottom=f"1px solid {C['border']}")),

            # Returns bar
            html.Div([
                html.Div("📅 How has the price changed?",
                         style=dict(fontWeight=700, fontSize="0.9rem", marginBottom="8px")),
                dcc.Graph(figure=chart_returns_bar(df, meta.get("name",ticker),
                                                   meta.get("color",C["accent"])),
                          config=dict(displayModeBar=False)),
            ], style=dict(padding="16px 24px")),
        ]
        main = html.Div(main_children, style=dict(flex=1, overflowY="auto"))

    return html.Div([sidebar, main],
                    style=dict(display="flex", flex=1, overflow="hidden",
                               minHeight="calc(100vh - 105px)"))


# ── TAB 3: ANALYTICS — Historical Data Explorer ───────────────────────────────
def _build_summary_table(tickers):
    rows = []
    for t in tickers:
        df  = load_df(t)
        sig = load_sig(t)
        meta = COMPANIES.get(t, {})
        if df is None: continue
        last = float(df["Close"].iloc[-1])
        rows.append({
            "Company":      meta.get("name", t),
            "Sector":       meta.get("sector", "—"),
            "Price (KES)":  f"{last:,.2f}",
            "1 Month":      f"{pct_change_over(df,21):+.1f}%",
            "3 Months":     f"{pct_change_over(df,63):+.1f}%",
            "6 Months":     f"{pct_change_over(df,126):+.1f}%",
            "1 Year":       f"{pct_change_over(df,252):+.1f}%",
            "All Time":     f"{pct_change_over(df,9999):+.1f}%",
            "Signal":    (sig.get("risk_adjusted_signal","—") if sig else "—"),
            "Source":    (("ML" if sig.get("signal_source","")=="ml" else "TA") if sig else "—"),
        })
    if not rows:
        return html.Div("No data available.", style=dict(color=C["muted"]))
    df_tbl = pd.DataFrame(rows)
    return dash_table.DataTable(
        data=df_tbl.to_dict("records"),
        columns=[{"name":c,"id":c} for c in df_tbl.columns],
        style_table=dict(overflowX="auto"),
        style_header=dict(background=C["panel"], color=C["muted"],
                          fontWeight=700, fontSize="0.75rem",
                          border=f"1px solid {C['border']}"),
        style_cell=dict(background=C["card"], color=C["text"],
                        fontSize="0.82rem", padding="8px 12px",
                        border=f"1px solid {C['border']}"),
        style_data_conditional=[
            {"if":{"filter_query":"{Signal} = BUY"},  "color":C["buy"],  "fontWeight":700},
            {"if":{"filter_query":"{Signal} = SELL"}, "color":C["sell"], "fontWeight":700},
            {"if":{"filter_query":"{Signal} = HOLD"}, "color":C["hold"], "fontWeight":700},
            {"if":{"filter_query":"{Source} = ML"},   "color":C["accent"]},
        ],
        tooltip_header={
            "Signal": "BUY = good time to invest · HOLD = keep what you have · SELL = consider exiting",
            "Source": "ML = AI/machine-learning model (more accurate) · TA = Technical Analysis rules (fallback when no ML data)",
        },
        tooltip_delay=0, tooltip_duration=None,
    )


def build_analytics(days=252, sector="All Companies", heatmap_ticker="SCOM.NR", custom_start=None, custom_end=None):
    tickers = SECTORS.get(sector, list(COMPANIES.keys()))

    summary_tbl = _build_summary_table(tickers)

    return html.Div([
        html.Div([
            html.Div("📊  Historical Data Explorer",
                     style=dict(fontSize="1.15rem", fontWeight=800, color=C["text"])),
            html.Div("How have these stocks performed historically? Compare growth, returns, and monthly patterns.",
                     style=dict(fontSize="0.83rem", color=C["muted"], marginTop="2px")),
        ], style=dict(padding="18px 24px 10px")),

        # Controls row
        html.Div([
            html.Div([
                html.Label("Period:", style=dict(color=C["muted"], fontSize="0.8rem",
                                                  marginBottom="4px")),
                dcc.Dropdown(id="analytics-period",
                             options=[{"label":k,"value":v} for k,v in PERIODS.items()],
                             value=days, clearable=False,
                             style=dict(background=C["card"], width="160px", fontSize="0.82rem")),
            ]),
            html.Div([
                html.Label("Sector:", style=dict(color=C["muted"], fontSize="0.8rem",
                                                 marginBottom="4px")),
                dcc.Dropdown(id="analytics-sector",
                             options=[{"label":k,"value":k} for k in SECTORS],
                             value=sector, clearable=False,
                             style=dict(background=C["card"], width="180px", fontSize="0.82rem")),
            ]),
            html.Div([
                html.Div([
                    html.Label("Companies:", style=dict(
                        color=C["muted"], fontSize="0.8rem",
                        marginRight="10px", alignSelf="center",
                    )),
                    html.Button("Select All", id="analytics-select-all", n_clicks=0, style=dict(
                        background="transparent", color=C["accent"],
                        border=f"1px solid {C['accent']}", borderRadius="12px",
                        padding="2px 10px", cursor="pointer", fontSize="0.72rem",
                        marginRight="6px",
                    )),
                    html.Button("Clear", id="analytics-clear", n_clicks=0, style=dict(
                        background="transparent", color=C["muted"],
                        border=f"1px solid {C['border']}", borderRadius="12px",
                        padding="2px 10px", cursor="pointer", fontSize="0.72rem",
                    )),
                ], style=dict(display="flex", alignItems="center", marginBottom="4px")),
                dcc.Dropdown(
                    id="analytics-companies",
                    options=[{"label": f"{COMPANIES[t]['short']} — {COMPANIES[t]['name']}", "value": t}
                             for t in tickers],
                    value=tickers,
                    multi=True,
                    clearable=True,
                    placeholder="Select companies to analyse…",
                    style=dict(background=C["card"], fontSize="0.82rem", minWidth="320px"),
                ),
            ], style=dict(flex="1")),
        ], style=dict(display="flex", gap="20px", flexWrap="wrap",
                      padding="0 24px 14px")),

        # Custom date range section — shown only when "Custom Range" is selected
        html.Div([
            html.Div([
                html.Label("Start Date", style=dict(color=C["muted"], fontSize="0.72rem", display="block", marginBottom="4px")),
                dcc.DatePickerSingle(
                    id="analytics-custom-start",
                    date=None,
                    display_format="DD MMM YYYY",
                    style=dict(fontSize="0.82rem"),
                ),
            ]),
            html.Div([
                html.Label("End Date", style=dict(color=C["muted"], fontSize="0.72rem", display="block", marginBottom="4px")),
                dcc.DatePickerSingle(
                    id="analytics-custom-end",
                    date=None,
                    display_format="DD MMM YYYY",
                    style=dict(fontSize="0.82rem"),
                ),
            ]),
            html.Button("Apply Range", id="analytics-apply-range", n_clicks=0, style=dict(
                background=C["accent"], color=C["header"],
                border="none", borderRadius="8px",
                padding="8px 20px", cursor="pointer", fontSize="0.82rem", fontWeight=700,
                alignSelf="flex-end",
            )),
        ], id="analytics-custom-range-section",
           style=dict(display="none", gap="14px", alignItems="flex-start", flexWrap="wrap")),

        html.Div(style=dict(borderBottom=f"1px solid {C['border']}", margin="0 0 0 0")),

        # Charts grid — all pre-populated with real data
        html.Div([
            # Row 1: indexed growth + ranked bar
            html.Div([
                html.Div([
                    html.Div("💰 Growth of KES 100 invested", style=dict(
                        fontWeight=700, fontSize="0.88rem", color=C["text"],
                        marginBottom="4px")),
                    html.Div("If you had invested KES 100 in each company at the start — who made you more money?",
                             style=dict(fontSize="0.72rem", color=C["muted"], marginBottom="8px")),
                    dcc.Loading(dcc.Graph(id="chart-indexed",
                              figure=chart_comparison_indexed(tickers, days),
                              config=dict(displayModeBar=False)),
                              type="circle", color=C["accent"]),
                ], style=dict(flex="2", background=C["card"],
                              border=f"1px solid {C['border']}", borderRadius="10px",
                              padding="16px")),
                html.Div([
                    html.Div("🏆 Best & Worst Performers", style=dict(
                        fontWeight=700, fontSize="0.88rem", color=C["text"],
                        marginBottom="4px")),
                    html.Div("Which company gained or lost the most?",
                             style=dict(fontSize="0.72rem", color=C["muted"], marginBottom="8px")),
                    dcc.Loading(dcc.Graph(id="chart-ranked",
                              figure=chart_performance_ranked(tickers, days),
                              config=dict(displayModeBar=False)),
                              type="circle", color=C["accent"]),
                ], style=dict(flex="1", background=C["card"],
                              border=f"1px solid {C['border']}", borderRadius="10px",
                              padding="16px")),
            ], style=dict(display="flex", gap="16px")),

            # Row 2: stacked price panels
            html.Div([
                html.Div("📉 Individual Price History (KES)", style=dict(
                    fontWeight=700, fontSize="0.88rem", color=C["text"],
                    marginBottom="4px")),
                html.Div("Actual share price in Kenyan Shillings for each company over the selected period.",
                         style=dict(fontSize="0.72rem", color=C["muted"], marginBottom="8px")),
                dcc.Loading(dcc.Graph(id="chart-all-prices",
                          figure=chart_all_prices(tickers, days),
                          config=dict(displayModeBar=True, scrollZoom=True)),
                          type="circle", color=C["accent"]),
            ], style=dict(background=C["card"], border=f"1px solid {C['border']}",
                          borderRadius="10px", padding="16px")),

            # Row 3: monthly heatmap
            html.Div([
                html.Div("📅 Monthly Returns Calendar", style=dict(
                    fontWeight=700, fontSize="0.88rem", color=C["text"],
                    marginBottom="4px")),
                html.Div("Green = the stock went up that month. Red = it went down. "
                         "The darker the colour, the bigger the move.",
                         style=dict(fontSize="0.72rem", color=C["muted"], marginBottom="12px")),
                html.Div([
                    dcc.Dropdown(
                        id="heatmap-ticker",
                        options=[{"label": f"{COMPANIES[t]['short']} — {COMPANIES[t]['name']}", "value": t}
                                 for t in sorted(tickers, key=lambda t: COMPANIES[t]["name"])],
                        value=heatmap_ticker,
                        clearable=False,
                        placeholder="Search or select a company…",
                        style=dict(flex="1", fontSize="0.82rem", minWidth="220px"),
                    ),
                    html.Button("Show Calendar", id="analytics-apply-heatmap", n_clicks=0,
                        style=dict(
                            background=C["accent"], color=C["header"],
                            border="none", borderRadius="8px",
                            padding="9px 22px", cursor="pointer",
                            fontSize="0.82rem", fontWeight=700, whiteSpace="nowrap",
                        )),
                ], style=dict(display="flex", gap="12px", alignItems="center",
                              marginBottom="12px", flexWrap="wrap")),
                dcc.Loading(dcc.Graph(id="chart-heatmap",
                          figure=chart_monthly_heatmap(heatmap_ticker),
                          config=dict(displayModeBar=False)),
                          type="circle", color=C["accent"]),
            ], style=dict(background=C["card"], border=f"1px solid {C['border']}",
                          borderRadius="10px", padding="16px")),

            # Row 4: performance summary table
            html.Div([
                html.Div("📋 Performance Summary — All Companies", style=dict(
                    fontWeight=700, fontSize="0.88rem", color=C["text"],
                    marginBottom="4px")),
                html.Div("Price changes across different time periods for every company we track.",
                         style=dict(fontSize="0.72rem", color=C["muted"], marginBottom="8px")),
                html.Div(id="summary-table", children=summary_tbl),
            ], style=dict(background=C["card"], border=f"1px solid {C['border']}",
                          borderRadius="10px", padding="16px")),

        ], style=dict(display="flex", flexDirection="column",
                      gap="16px", padding="16px 24px 24px")),
    ], style=dict(overflowY="auto", maxHeight="calc(100vh - 105px)"))


# ── TAB 4: IMPORT & ANALYSE — Add New Company Data ───────────────────────────
def build_import_tab():
    return html.Div([
        html.Div([
            html.Div("📥  Import & Analyse New Company Data", style=dict(
                fontSize="1.15rem", fontWeight=800, color=C["text"])),
            html.Div("Upload a CSV of daily prices, give the company a ticker name, then choose how to analyse it.",
                     style=dict(fontSize="0.83rem", color=C["muted"], marginTop="2px")),
        ], style=dict(padding="20px 24px 16px",
                      borderBottom=f"1px solid {C['border']}")),

        html.Div([
            # Upload zone
            dcc.Upload(id="csv-upload-tab",
                accept=".csv",
                children=html.Div([
                    html.Div("📂", style=dict(fontSize="2.5rem", marginBottom="8px")),
                    html.Div("Click to upload a CSV file", style=dict(
                        fontWeight=700, color=C["accent"], fontSize="1rem")),
                    html.Div("or drag and drop here", style=dict(
                        color=C["muted"], fontSize="0.82rem")),
                    html.Div("Required columns: Date, Open, High, Low, Close, Volume",
                             style=dict(color=C["muted"], fontSize="0.75rem", marginTop="8px")),
                    html.Div(id="upload-filename", style=dict(
                        color=C["buy"], fontSize="0.78rem", marginTop="8px")),
                ], style=dict(textAlign="center", padding="40px")),
                style=dict(border=f"2px dashed {C['border']}", borderRadius="12px",
                           background=C["card"], cursor="pointer", marginBottom="16px",
                           transition="border 0.2s"),
            ),

            # Ticker input
            html.Div([
                html.Label("Company ticker / name:", style=dict(
                    color=C["muted"], fontSize="0.82rem", marginBottom="6px",
                    display="block")),
                dcc.Input(id="import-tab-name", placeholder="e.g. BAMB.NR or UCHUMI.NR",
                          style=dict(background=C["card"], border=f"1px solid {C['border']}",
                                     color=C["text"], borderRadius="8px",
                                     padding="8px 14px", width="260px",
                                     fontSize="0.85rem")),
            ], style=dict(marginBottom="16px")),

            # Buttons row
            html.Div([
                html.Button("⚡ Quick Analysis (30 sec)", id="run-quick-import-btn", style=dict(
                    background=C["accent"], color=C["header"], fontWeight=700,
                    border="none", borderRadius="8px",
                    padding="10px 22px", cursor="pointer", fontSize="0.88rem")),
                html.Button("🧠 Full AI Analysis (10–15 min)", id="run-full-import-btn", style=dict(
                    background="transparent", color=C["accent"], fontWeight=700,
                    border=f"2px solid {C['accent']}", borderRadius="8px",
                    padding="10px 22px", cursor="pointer", fontSize="0.88rem")),
            ], style=dict(display="flex", gap="12px", flexWrap="wrap", marginBottom="16px")),

            # Status display
            html.Div([
                dcc.Loading(html.Div(id="import-tab-status"), type="circle", color=C["accent"]),
                html.Div(id="pipeline-status-display", style=dict(
                    marginTop="8px", padding="12px 16px",
                    background=C["card"], border=f"1px solid {C['border']}",
                    borderRadius="8px", display="none",
                )),
            ], style=dict(marginBottom="16px")),

            # How-to instructions
            html.Div([
                html.Div("📌 How to get NSE data:", style=dict(
                    fontWeight=700, color=C["text"], marginBottom="8px", fontSize="0.88rem")),
                html.Ol([
                    html.Li("Go to the NSE website: nse.co.ke",
                            style=dict(fontSize="0.82rem", color=C["muted"], marginBottom="4px")),
                    html.Li("Navigate to Trade Statistics → Historical Data",
                            style=dict(fontSize="0.82rem", color=C["muted"], marginBottom="4px")),
                    html.Li("Select your company and date range",
                            style=dict(fontSize="0.82rem", color=C["muted"], marginBottom="4px")),
                    html.Li("Download the CSV file",
                            style=dict(fontSize="0.82rem", color=C["muted"], marginBottom="4px")),
                    html.Li("Upload it above, enter the ticker, then click Quick or Full Analysis",
                            style=dict(fontSize="0.82rem", color=C["muted"])),
                ]),
                html.Div([
                    html.Div("Quick Analysis vs Full AI Analysis:", style=dict(
                        fontWeight=700, color=C["text"], marginTop="14px", marginBottom="6px",
                        fontSize="0.85rem")),
                    html.Div([
                        html.Span("⚡ Quick (30 sec)", style=dict(color=C["accent"], fontWeight=700)),
                        html.Span(" — Runs XGBoost only. Gets you a fast signal with no waiting.",
                                  style=dict(color=C["muted"], fontSize="0.8rem")),
                    ], style=dict(marginBottom="4px")),
                    html.Div([
                        html.Span("🧠 Full AI (10–15 min)", style=dict(color=C["accent"], fontWeight=700)),
                        html.Span(" — Runs the complete pipeline: ARIMA + LSTM + XGBoost ensemble. "
                                  "Much more accurate but takes longer. Runs in the background.",
                                  style=dict(color=C["muted"], fontSize="0.8rem")),
                    ]),
                ]),
            ], style=dict(background=C["card"], border=f"1px solid {C['border']}",
                          borderRadius="10px", padding="16px")),

        ], style=dict(padding="16px 24px 24px", maxWidth="720px")),
    ], style=dict(overflowY="auto", maxHeight="calc(100vh - 105px)"))


# ── Callbacks: Company tab ────────────────────────────────────────────────────
@app.callback(
    Output("price-chart-company","figure"),
    Input("price-period","value"),
    State("selected-ticker","data"),
    prevent_initial_call=True,
)
def update_price_chart(days, ticker):
    df = load_df(ticker)
    if df is None: return go.Figure()
    return chart_price_simple(df, ticker, days)


@app.callback(
    Output("analysis-store","data"),
    Output("quick-status","children"),
    Input("quick-btn","n_clicks"),
    State("selected-ticker","data"),
    State("analysis-store","data"),
    prevent_initial_call=True,
)
def run_quick(n, ticker, store):
    if not n: return store, ""
    df = load_df(ticker)
    if df is None:
        return store, html.Span("No data found.", style=dict(color=C["sell"]))
    try:
        result = run_quick_fn(df, ticker)
        store = store or {}
        store[ticker] = {k:v for k,v in result.items()
                         if k not in ("ma_df",) and not isinstance(v, pd.DataFrame)}
        return store, html.Span("✓ Done", style=dict(color=C["buy"], fontSize="0.8rem"))
    except Exception as e:
        return store, html.Span(f"Error: {e}", style=dict(color=C["sell"], fontSize="0.75rem"))


# ── Callbacks: Analytics tab ─────────────────────────────────────────────────
@app.callback(
    Output("analytics-companies","options"),
    Output("analytics-companies","value"),
    Output("heatmap-ticker","options"),
    Input("analytics-sector","value"),
    prevent_initial_call=True,
)
def update_sector_filters(sector):
    tickers = SECTORS.get(sector, list(COMPANIES.keys()))
    company_opts = [{"label": f"{COMPANIES[t]['short']} — {COMPANIES[t]['name']}", "value": t}
                    for t in tickers]
    heatmap_opts = [{"label": f"{COMPANIES[t]['short']} — {COMPANIES[t]['name']}", "value": t}
                    for t in sorted(tickers, key=lambda t: COMPANIES[t]["name"])]
    return company_opts, tickers, heatmap_opts


@app.callback(
    Output("analytics-companies", "value", allow_duplicate=True),
    Input("analytics-select-all", "n_clicks"),
    Input("analytics-clear", "n_clicks"),
    State("analytics-companies", "options"),
    State("analytics-sector", "value"),
    prevent_initial_call=True,
)
def analytics_bulk_select(sel_clicks, clr_clicks, current_opts, sector):
    from dash import ctx
    if ctx.triggered_id == "analytics-clear":
        return []
    # Select All — return all tickers currently visible (filtered by sector)
    return [o["value"] for o in (current_opts or [])]


@app.callback(
    Output("chart-indexed","figure"),
    Output("chart-ranked","figure"),
    Output("chart-all-prices","figure"),
    Output("summary-table","children"),
    Input("analytics-period","value"),
    Input("analytics-sector","value"),
    Input("analytics-companies","value"),
    Input("analytics-apply-range","n_clicks"),
    State("analytics-custom-start","date"),
    State("analytics-custom-end","date"),
    prevent_initial_call=True,
)
def update_analytics_charts(days, sector, companies, n_apply, custom_start, custom_end):
    tickers = companies if companies else SECTORS.get(sector, list(COMPANIES.keys()))
    use_days = days if days != 0 else 9999
    use_start = custom_start if days == 0 else None
    use_end   = custom_end   if days == 0 else None
    return (
        chart_comparison_indexed(tickers, use_days, use_start, use_end),
        chart_performance_ranked(tickers, use_days, use_start, use_end),
        chart_all_prices(tickers, use_days, use_start, use_end),
        _build_summary_table(tickers),
    )


@app.callback(
    Output("chart-heatmap","figure"),
    Input("analytics-apply-heatmap","n_clicks"),
    State("heatmap-ticker","value"),
    prevent_initial_call=True,
)
def update_heatmap(_n, ticker):
    if not ticker:
        return dash.no_update
    return chart_monthly_heatmap(ticker)


# ── Callbacks: Import tab — upload feedback ───────────────────────────────────
@app.callback(
    Output("upload-filename","children"),
    Input("csv-upload-tab","filename"),
    prevent_initial_call=True,
)
def show_upload_name(filename):
    if not filename:
        return ""
    return f"✓ File ready: {filename}"


# ── Callbacks: Import tab — Quick Analysis button ────────────────────────────
@app.callback(
    Output("selected-ticker","data",        allow_duplicate=True),
    Output("analysis-store","data",         allow_duplicate=True),
    Output("import-tab-status","children"),
    Output("main-tabs","value",             allow_duplicate=True),
    Input("run-quick-import-btn","n_clicks"),
    State("import-tab-name","value"),
    State("csv-upload-tab","contents"),
    State("analysis-store","data"),
    prevent_initial_call=True,
)
def process_quick_import(n, name, contents, store):
    return _process_csv_quick(n, name, contents, store)


def _process_csv_quick(n, name, contents, store):
    no = (dash.no_update, store or {},
          html.Span("Please fill in the company name and upload a CSV.",
                    style=dict(color=C["hold"])),
          dash.no_update)
    if not n or not name or not contents:
        return no
    ticker = name.strip().upper()
    if not ticker.endswith(".NR"):
        ticker += ".NR"
    try:
        _, cs = contents.split(",")
        decoded = base64.b64decode(cs)
        from src.data.fetcher import load_from_csv
        from src.data.cleaner import clean_ohlcv, save_cleaned
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="wb") as f:
            f.write(decoded)
            tmp = f.name
        raw = load_from_csv(tmp, ticker=ticker)
        os.unlink(tmp)
        cleaned, _ = clean_ohlcv(raw, ticker=ticker)
        save_cleaned(cleaned, ticker)
        result = run_quick_fn(cleaned, ticker)
        store = store or {}
        store[ticker] = {k:v for k,v in result.items()
                         if k not in ("ma_df",) and not isinstance(v, pd.DataFrame)}
        return (ticker, store,
                html.Span(f"✓ {ticker} analysed! Signal: {result.get('risk_adjusted_signal','—')}",
                          style=dict(color=C["buy"])),
                "company")
    except Exception as e:
        err = str(e).lower()
        if any(w in err for w in ("column", "columns", "keyerror")):
            msg = "CSV must have columns: Date, Open, High, Low, Close, Volume. Please check your file."
        elif any(w in err for w in ("date", "parse", "time")):
            msg = "Could not read dates. Ensure the Date column uses YYYY-MM-DD format."
        elif "empty" in err or "no data" in err:
            msg = "The file appears to be empty. Please upload a valid CSV with price data."
        else:
            msg = "Import failed. Please check your CSV format and try again."
        return (dash.no_update, store or {},
                html.Span(msg, style=dict(color=C["sell"], fontSize="0.75rem")),
                dash.no_update)


# ── Callbacks: Import tab — Full AI Analysis (background subprocess) ──────────
@app.callback(
    Output("pipeline-ticker","data"),
    Output("pipeline-csv-path","data"),
    Output("pipeline-status-display","children"),
    Output("pipeline-status-display","style"),
    Output("pipeline-poll","disabled"),
    Input("run-full-import-btn","n_clicks"),
    State("import-tab-name","value"),
    State("csv-upload-tab","contents"),
    State("csv-upload-tab","filename"),
    prevent_initial_call=True,
)
def start_full_pipeline(n, name, contents, filename):
    _hidden = dict(display="none")
    _visible = dict(
        marginTop="8px", padding="12px 16px",
        background=C["card"], border=f"1px solid {C['border']}",
        borderRadius="8px", display="block",
    )
    no = ("", "", dash.no_update, _hidden, True)
    if not n or not name or not contents:
        return ("", "",
                html.Span("Please enter a ticker and upload a CSV first.",
                          style=dict(color=C["hold"])),
                _visible, True)

    ticker = name.strip().upper()
    if not ticker.endswith(".NR"):
        ticker += ".NR"

    try:
        # Save CSV to DATA_RAW
        _, cs = contents.split(",")
        decoded = base64.b64decode(cs)
        DATA_RAW.mkdir(parents=True, exist_ok=True)
        safe_name = filename or f"{ticker.replace('.','_')}.csv"
        csv_path = DATA_RAW / safe_name
        with open(csv_path, "wb") as f:
            f.write(decoded)

        # Write initial status JSON
        DATA_FEATURES.mkdir(parents=True, exist_ok=True)
        status_path = DATA_FEATURES / f"{ticker.replace('.','_')}_pipeline_status.json"
        with open(status_path, "w") as f:
            json.dump({"status": "running", "step": "Saving data..."}, f)

        # Launch subprocess
        subprocess.Popen(
            [sys.executable, str(Path(__file__).parent / "main.py"),
             "--ticker", ticker, "--csv", str(csv_path)],
            cwd=str(Path(__file__).parent),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        msg = html.Div([
            html.Div("🧠 Full AI pipeline started!", style=dict(
                color=C["accent"], fontWeight=700, marginBottom="4px")),
            html.Div(f"Analysing {ticker} — this runs ARIMA + LSTM + XGBoost in the background.",
                     style=dict(color=C["muted"], fontSize="0.82rem", marginBottom="4px")),
            html.Div("Status will update every 3 seconds below. You can use other tabs while waiting.",
                     style=dict(color=C["muted"], fontSize="0.78rem")),
        ])
        return (ticker, str(csv_path), msg, _visible, False)

    except Exception as e:
        return ("", "",
                html.Span(f"Error starting pipeline: {e}",
                          style=dict(color=C["sell"], fontSize="0.75rem")),
                _visible, True)


@app.callback(
    Output("pipeline-status-display","children",  allow_duplicate=True),
    Output("pipeline-status-display","style",     allow_duplicate=True),
    Output("pipeline-poll","disabled",             allow_duplicate=True),
    Output("selected-ticker","data",               allow_duplicate=True),
    Output("main-tabs","value",                    allow_duplicate=True),
    Input("pipeline-poll","n_intervals"),
    State("pipeline-ticker","data"),
    State("selected-ticker","data"),
    prevent_initial_call=True,
)
def poll_pipeline_status(n_intervals, pipeline_ticker, current_ticker):
    _hidden = dict(display="none")
    _visible = dict(
        marginTop="8px", padding="12px 16px",
        background=C["card"], border=f"1px solid {C['border']}",
        borderRadius="8px", display="block",
    )
    no_nav = (dash.no_update, dash.no_update, dash.no_update,
              current_ticker, dash.no_update)

    if not pipeline_ticker:
        return (dash.no_update, _hidden, True, current_ticker, dash.no_update)

    status_path = DATA_FEATURES / f"{pipeline_ticker.replace('.','_')}_pipeline_status.json"
    signal_path = DATA_FEATURES / f"{pipeline_ticker.replace('.','_')}_signal.json"

    # Check if full signal file appeared (pipeline completed)
    if signal_path.exists():
        sig = load_sig(pipeline_ticker)
        final_signal = sig.get("risk_adjusted_signal", "—") if sig else "—"
        adv = ADVICE.get(final_signal, {})
        done_msg = html.Div([
            html.Div(f"✅ Done! Analysis complete for {pipeline_ticker}",
                     style=dict(color=C["buy"], fontWeight=700, marginBottom="6px")),
            html.Div([
                html.Span("Signal: ", style=dict(color=C["muted"])),
                html.Span(f"{adv.get('emoji','')} {final_signal}",
                          style=dict(color=adv.get("color", C["text"]),
                                     fontWeight=700, fontSize="1.1rem")),
            ]),
            html.Div("Switching to Company tab…",
                     style=dict(color=C["muted"], fontSize="0.78rem", marginTop="4px")),
        ])
        # Stop polling, navigate to company tab
        return (done_msg, _visible, True, pipeline_ticker, "company")

    # Read status JSON if available
    step = "Training models…"
    if status_path.exists():
        try:
            with open(status_path) as f:
                st = json.load(f)
            step = st.get("step", "Training models…")
        except Exception:
            pass

    steps = ["Saving data…", "Training models…", "Running ARIMA…",
             "Running LSTM…", "Running XGBoost…", "Generating signal…"]
    progress_msg = html.Div([
        html.Div(f"🔄 Running full pipeline for {pipeline_ticker}…",
                 style=dict(color=C["accent"], fontWeight=700, marginBottom="4px")),
        html.Div(f"Current step: {step}",
                 style=dict(color=C["muted"], fontSize="0.82rem")),
        html.Div("This may take 10–15 minutes. Status updates every 3 seconds.",
                 style=dict(color=C["muted"], fontSize="0.75rem", marginTop="4px")),
    ])
    return (progress_msg, _visible, False, current_ticker, dash.no_update)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — NSE DATA EXPLORER
# ══════════════════════════════════════════════════════════════════════════════

_ARCHIVE_DIR = Path(r"C:\Users\moeng\Downloads\archive")

_NSE_COLS_ORDERED = [
    "Date", "Code", "Name",
    "12m Low", "12m High",
    "Day Low", "Day High", "Day Price",
    "Previous", "Change", "Change%",
    "Volume", "Adjusted Price",
]

_FIELD_GLOSSARY = {
    "Date":           "Trading date",
    "Code":           "NSE ticker symbol (e.g. EQTY, KCB, SCOM)",
    "Name":           "Full company name as listed on the NSE",
    "12m Low":        "Lowest price traded in the past 12 months",
    "12m High":       "Highest price traded in the past 12 months",
    "Day Low":        "Lowest price traded during this session",
    "Day High":       "Highest price traded during this session",
    "Day Price":      "Official closing price for the session (KES)",
    "Previous":       "Previous session's official closing price (KES)",
    "Change":         "Price change in KES  (Day Price − Previous)",
    "Change%":        "Percentage price change vs previous session",
    "Volume":         "Number of shares traded during the session",
    "Adjusted Price": "Price adjusted for corporate actions (splits, dividends)",
}

_ALL_NSE_CODES = [
    "ABSA","ALP","AMAC","BAT","BKG","BOC","BRIT","CARB","CGEN","CIC",
    "COOP","CRWN","CTUM","DTK","EABL","EGAD","EQTY","EVRD","FMLY","FTGH",
    "GLD","HAFR","HFCK","IMH","JUB","KAPC","KCB","KEGN","KNRE","KPC",
    "KPLC","KQ","KUKZ","KURV","LBTY","LIMT","LKL","NBV","NCBA","NMG",
    "NSE","OCH","PORT","SASN","SBIC","SCAN","SCBK","SCOM","SGL","SKL",
    "SLAM","SMER","SMWF","TOTL","TPSE","TRFC","UCHM","UMME","UNGA","WTK","XPRS",
]

# Company name lookup (code → full name) built once from the latest archive year
def _build_name_map():
    path = _ARCHIVE_DIR / "NSE_data_all_stocks_2026.csv"
    if not path.exists():
        return {}
    try:
        df = pd.read_csv(path, dtype=str, usecols=["Code", "Name"])
        df["Code"] = df["Code"].str.strip()
        df["Name"] = df["Name"].str.strip()
        return dict(zip(df["Code"], df["Name"]))
    except Exception:
        return {}

_NAME_MAP = _build_name_map()

# Earliest / latest dates available in the archive (used for picker bounds)
def _archive_date_bounds():
    years = sorted(
        int(p.stem.split("_")[-1])
        for p in _ARCHIVE_DIR.glob("NSE_data_all_stocks_????.csv")
        if p.stem.split("_")[-1].isdigit()
    )
    min_year = years[0] if years else 2000
    max_year = years[-1] if years else 2026
    return f"{min_year}-01-01", f"{max_year}-12-31"

_ARCHIVE_MIN_DATE, _ARCHIVE_MAX_DATE = _archive_date_bounds()


# ── Archive data loader ───────────────────────────────────────────────────────

def _load_archive_range(codes, start_date, end_date):
    start = pd.to_datetime(start_date)
    end   = pd.to_datetime(end_date)
    years = range(start.year, end.year + 1)
    frames = []

    def _load_one(path):
        try:
            raw = pd.read_csv(path, dtype=str)
            df  = _normalise_archive_cols(raw)
            if "Code" not in df.columns or "Date" not in df.columns:
                return
            df["Code"] = df["Code"].str.strip()
            if codes:
                df = df[df["Code"].isin(codes)]
            if df.empty:
                return
            df["_dt"] = pd.to_datetime(
                df["Date"].str.strip(), dayfirst=False, format="mixed", errors="coerce"
            )
            df = df.dropna(subset=["_dt"])
            frames.append(df)
        except Exception:
            pass

    for year in years:
        _load_one(_ARCHIVE_DIR / f"NSE_data_all_stocks_{year}.csv")

    # Load patch files whose dates fall in the requested range
    for patch in sorted(_ARCHIVE_DIR.glob("NSE_patch_*.csv")):
        try:
            patch_date = pd.to_datetime(patch.stem.replace("NSE_patch_", ""))
            if start <= patch_date <= end:
                _load_one(patch)
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    combined = combined[(combined["_dt"] >= start) & (combined["_dt"] <= end)]
    combined = combined.sort_values(["_dt", "Code"]).reset_index(drop=True)

    # Repair decimal-point-dropped prices (e.g. 3540 → 35.40) per company per price column
    _PRICE_REPAIR_COLS = [
        c for c in ["Day Price", "Day High", "Day Low", "Previous", "12m Low", "12m High"]
        if c in combined.columns
    ]
    for col in _PRICE_REPAIR_COLS:
        numeric = pd.to_numeric(
            combined[col].astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce",
        )
        repaired = numeric.copy()
        for code in combined["Code"].unique():
            mask = combined["Code"] == code
            sub_sorted = combined.loc[mask].sort_values("_dt")
            s = numeric.loc[sub_sorted.index]
            if s.notna().sum() >= 3:
                repaired.loc[sub_sorted.index] = _repair_decimal_errors(s).values
        changed = (repaired - numeric).abs() > 0.01
        if changed.any():
            combined.loc[changed, col] = repaired[changed].round(2).astype(str)

    return combined


def _to_num(series):
    return pd.to_numeric(
        series.astype(str)
              .str.replace(",", "", regex=False)
              .str.replace("%", "", regex=False)
              .str.strip()
              .replace(["-", "", "nan", "NaN", "N/A", "0"], None),
        errors="coerce",
    )


# ── Stats & Top-Movers helpers ────────────────────────────────────────────────

def _compute_stats(df):
    """Return summary stats dict and top-movers list from loaded archive df."""
    if df.empty:
        return {}, []

    prices = _to_num(df["Day Price"]) if "Day Price" in df.columns else pd.Series(dtype=float)
    changes = _to_num(df["Change%"]) if "Change%" in df.columns else pd.Series(dtype=float)
    volumes = _to_num(df["Volume"])  if "Volume"  in df.columns else pd.Series(dtype=float)

    # Per-company total return: last Day Price / first Day Price - 1
    # Exclude stocks where price scale looks inconsistent (ratio > 100x suggests data error)
    returns = {}
    for code, grp in df.groupby("Code"):
        p = _to_num(grp.sort_values("_dt")["Day Price"]).dropna()
        if len(p) >= 2 and p.min() > 0:
            ratio = p.max() / p.min()
            if ratio < 10:   # only include if max/min < 10x (sane price range)
                returns[code] = (p.iloc[-1] / p.iloc[0] - 1) * 100

    best_code  = max(returns, key=returns.get) if returns else "—"
    worst_code = min(returns, key=returns.get) if returns else "—"
    best_pct   = returns.get(best_code, 0)
    worst_pct  = returns.get(worst_code, 0)

    biggest_day = changes.abs().max()
    biggest_row = df.loc[changes.abs().idxmax()] if not changes.dropna().empty else None

    stats = {
        "best_code":  best_code,
        "best_pct":   best_pct,
        "worst_code": worst_code,
        "worst_pct":  worst_pct,
        "biggest_day_pct": biggest_day if pd.notna(biggest_day) else 0,
        "biggest_day_info": (
            f"{biggest_row['Code']} on {pd.to_datetime(biggest_row['_dt']).strftime('%d %b %Y')}"
            if biggest_row is not None else "—"
        ),
        "avg_volume":  volumes.mean(),
        "total_records": len(df),
    }

    # Top movers — only meaningful with ≥2 companies; deduplicate gainers/losers
    sorted_returns = sorted(returns.items(), key=lambda x: x[1], reverse=True)
    if len(returns) <= 1:
        top_gainers, top_losers = [], []
    else:
        top_gainers = [(c, p) for c, p in sorted_returns[:3]  if p > 0]
        seen = {c for c, _ in top_gainers}
        top_losers  = [(c, p) for c, p in reversed(sorted_returns) if p < 0 and c not in seen][:3]
    movers = [("gain", c, p) for c, p in top_gainers] + [("loss", c, p) for c, p in top_losers]
    return stats, movers


# ── Chart builder ─────────────────────────────────────────────────────────────

def _build_chart(df, chart_type="line"):
    if df.empty:
        return go.Figure()

    palette = ["#38bdf8","#a78bfa","#f472b6","#fb923c","#34d399",
               "#f59e0b","#60a5fa","#e879f9","#4ade80","#fbbf24",
               "#c084fc","#818cf8","#f87171","#2dd4bf","#facc15"]

    codes = sorted(df["Code"].unique())
    multi = len(codes) > 1

    if chart_type == "candlestick" and not multi:
        # Single-stock candlestick + volume subplot
        code = codes[0]
        sub  = df[df["Code"] == code].sort_values("_dt")
        close  = _to_num(sub["Day Price"])
        open_  = _to_num(sub["Previous"])
        high   = _to_num(sub["Day High"])   if "Day High" in sub.columns else close
        low    = _to_num(sub["Day Low"])    if "Day Low"  in sub.columns else close
        volume = _to_num(sub["Volume"])     if "Volume"   in sub.columns else pd.Series(dtype=float)
        name   = _NAME_MAP.get(code, code)

        fig = make_subplots(
            rows=2, cols=1, shared_xaxes=True,
            row_heights=[0.72, 0.28], vertical_spacing=0.03,
        )
        vol_colors = [C["buy"] if (c >= o) else C["sell"]
                      for c, o in zip(close.fillna(0), open_.fillna(0))]
        fig.add_trace(go.Candlestick(
            x=sub["_dt"], open=open_, high=high, low=low, close=close,
            name=f"{code} — {name}",
            increasing_line_color=C["buy"], decreasing_line_color=C["sell"],
        ), row=1, col=1)
        fig.add_trace(go.Bar(
            x=sub["_dt"], y=volume, name="Volume",
            marker_color=vol_colors, opacity=0.7, showlegend=False,
        ), row=2, col=1)
        fig.update_layout(
            **CHART_BASE, height=480,
            title=dict(text=f"{code}  —  {name}  |  Candlestick + Volume", font=dict(color=C["accent"])),
            xaxis_rangeslider_visible=False,
        )
        fig.update_yaxes(title_text="Price (KES)", gridcolor=C["border"], row=1, col=1)
        fig.update_yaxes(title_text="Volume",      gridcolor=C["border"], row=2, col=1)
        fig.update_xaxes(gridcolor=C["border"], zeroline=False)
        return fig

    # Line chart (multi-stock or single-stock)
    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True,
        row_heights=[0.72, 0.28], vertical_spacing=0.03,
    )
    for i, code in enumerate(codes):
        sub    = df[df["Code"] == code].sort_values("_dt")
        price  = _to_num(sub["Day Price"])
        volume = _to_num(sub["Volume"]) if "Volume" in sub.columns else pd.Series(dtype=float)
        name   = _NAME_MAP.get(code, code)
        color  = palette[i % len(palette)]

        fig.add_trace(go.Scatter(
            x=sub["_dt"], y=price,
            name=f"{code} — {name}" if not multi else code,
            line=dict(color=color, width=1.8),
            hovertemplate=f"<b>{code}</b><br>%{{x|%d %b %Y}}<br>KES %{{y:,.2f}}<extra></extra>",
        ), row=1, col=1)

        if not multi:
            vol_colors = [C["buy"] if (c >= p) else C["sell"]
                          for c, p in zip(
                              _to_num(sub["Day Price"]).fillna(0),
                              _to_num(sub["Previous"]).fillna(0),
                          )]
            fig.add_trace(go.Bar(
                x=sub["_dt"], y=volume, name="Volume",
                marker_color=vol_colors, opacity=0.7, showlegend=False,
            ), row=2, col=1)

    title_text = (
        "Daily Price — Day Price (KES)" if multi
        else f"{codes[0]}  —  {_NAME_MAP.get(codes[0], codes[0])}  |  Price + Volume"
    )
    fig.update_layout(
        **{**CHART_BASE, "legend": dict(bgcolor="rgba(0,0,0,0)", font=dict(size=10))},
        height=480,
        title=dict(text=title_text, font=dict(color=C["accent"])),
        hovermode="x unified",
    )
    fig.update_yaxes(title_text="Price (KES)", gridcolor=C["border"], row=1, col=1)
    fig.update_yaxes(title_text="Volume",      gridcolor=C["border"], row=2, col=1)
    fig.update_xaxes(gridcolor=C["border"], zeroline=False)
    return fig


# ── UI helpers ────────────────────────────────────────────────────────────────

def _stat_card(label, value, value_color=None, sub=None):
    return html.Div([
        html.Div(label, style=dict(
            fontSize="0.62rem", color=C["muted"],
            textTransform="uppercase", letterSpacing="0.08em", marginBottom="4px",
        )),
        html.Div(value, style=dict(
            fontSize="1.1rem", fontWeight=800,
            color=value_color or C["text"],
        )),
        html.Div(sub or "", style=dict(fontSize="0.68rem", color=C["muted"], marginTop="2px")),
    ], style=dict(
        background=C["card"], border=f"1px solid {C['border']}",
        borderRadius="10px", padding="14px 18px", flex="1", minWidth="160px",
    ))


def _mover_badge(code, pct, is_gain):
    color = C["buy"] if is_gain else C["sell"]
    bg    = "#052e1640" if is_gain else "#2d0a0a40"
    arrow = "▲" if is_gain else "▼"
    name  = _NAME_MAP.get(code, code)
    return html.Div([
        html.Span(f"{arrow} {code}", style=dict(fontWeight=800, color=color, fontSize="0.82rem")),
        html.Div(name,               style=dict(fontSize="0.65rem", color=C["muted"])),
        html.Div(f"{pct:+.1f}%",    style=dict(fontWeight=700, color=color, fontSize="0.85rem")),
    ], style=dict(
        background=bg, border=f"1px solid {color}50",
        borderRadius="8px", padding="8px 14px", textAlign="center", minWidth="90px",
    ))


def _build_interpretation(df, stats, movers):
    """Generate a natural-language analysis panel for the explorer results."""
    n_codes = df["Code"].nunique()
    returns_by_code: dict[str, float] = {}
    volatility_by_code: dict[str, float] = {}

    for code, grp in df.groupby("Code"):
        prices  = _to_num(grp.sort_values("_dt")["Day Price"]).dropna()
        changes = _to_num(grp.sort_values("_dt").get("Change%", pd.Series(dtype=float))).dropna()
        if len(prices) >= 2 and prices.min() > 0:
            ratio = prices.max() / prices.min()
            if ratio < 10:
                returns_by_code[code] = (prices.iloc[-1] / prices.iloc[0] - 1) * 100
        if len(changes) >= 2:
            volatility_by_code[code] = float(changes.std())

    if not returns_by_code:
        return html.Div()

    if n_codes == 1:
        code  = next(iter(returns_by_code))
        ret   = returns_by_code[code]
        name  = _NAME_MAP.get(code, code)
        vol   = volatility_by_code.get(code, 0)
        sector = COMPANIES.get(code, {}).get("sector", "")

        grp     = df[df["Code"] == code].sort_values("_dt")
        prices  = _to_num(grp["Day Price"]).dropna()
        changes = _to_num(grp.get("Change%", pd.Series(dtype=float))).dropna()
        n_days  = len(prices)

        # Trend: second half vs first half average
        mid  = max(1, n_days // 2)
        trend = (
            "strengthening" if prices.iloc[mid:].mean() > prices.iloc[:mid].mean() * 1.02
            else "weakening"  if prices.iloc[mid:].mean() < prices.iloc[:mid].mean() * 0.98
            else "moving sideways"
        )

        if ret >= 5:
            sentiment, badge_color = "Bullish", C["buy"]
        elif ret <= -5:
            sentiment, badge_color = "Bearish", C["sell"]
        else:
            sentiment, badge_color = "Neutral", C["hold"]

        vol_label = "high" if vol > 2 else "moderate" if vol > 0.8 else "low"
        pos_days  = int((changes > 0).sum())
        pct_pos   = pos_days / len(changes) * 100 if len(changes) else 0

        lines = [
            f"{name} ({code}){' — ' + sector if sector else ''} returned "
            f"<b style='color:{badge_color}'>{ret:+.1f}%</b> over the selected period.",
            f"Price momentum is <b>{trend}</b>, with <b>{vol_label} volatility</b> "
            f"(daily swing avg: {vol:.2f}%).",
            f"The stock closed higher on <b>{pos_days} of {len(changes)} sessions</b> "
            f"({pct_pos:.0f}% positive days).",
        ]
        if stats.get("biggest_day_pct"):
            lines.append(
                f"Largest single-day move: <b>{stats['biggest_day_pct']:+.1f}%</b> "
                f"({stats['biggest_day_info']})."
            )
        if ret >= 5:
            lines.append("Momentum is positive — consider reviewing the signal on the Analytics tab.")
        elif ret <= -5:
            lines.append("Sustained decline detected — review fundamentals before adding to position.")
        else:
            lines.append("The stock is ranging; no clear directional bias from this period.")

    else:
        n_pos   = sum(1 for v in returns_by_code.values() if v > 0)
        pct_pos = n_pos / len(returns_by_code) * 100
        avg_ret = sum(returns_by_code.values()) / len(returns_by_code)

        if pct_pos >= 60:
            sentiment, badge_color = "Bullish", C["buy"]
        elif pct_pos <= 40:
            sentiment, badge_color = "Bearish", C["sell"]
        else:
            sentiment, badge_color = "Mixed", C["hold"]

        # Sector breakdown
        sector_rets: dict[str, list[float]] = {}
        for code, ret in returns_by_code.items():
            sec = COMPANIES.get(code, {}).get("sector", "Other")
            sector_rets.setdefault(sec, []).append(ret)
        sector_avg = {s: sum(v) / len(v) for s, v in sector_rets.items() if v}

        buy_color = C["buy"]
        lines = [
            f"Across <b>{n_codes} companies</b>, <b style='color:{buy_color}'>{n_pos} ({pct_pos:.0f}%)</b> "
            f"showed positive returns — market tone is <b>{sentiment.lower()}</b>.",
            f"Average return: <b style='color:{badge_color}'>{avg_ret:+.1f}%</b>.",
        ]

        if sector_avg:
            best_s  = max(sector_avg, key=sector_avg.get)
            worst_s = min(sector_avg, key=sector_avg.get)
            if best_s != worst_s:
                lines.append(
                    f"Best sector: <b>{best_s}</b> (avg {sector_avg[best_s]:+.1f}%) · "
                    f"Weakest: <b>{worst_s}</b> (avg {sector_avg[worst_s]:+.1f}%)."
                )

        gainers = [(c, p) for t, c, p in movers if t == "gain"]
        losers  = [(c, p) for t, c, p in movers if t == "loss"]
        if gainers:
            g_str = ", ".join(f"<b>{c}</b> ({p:+.1f}%)" for c, p in gainers)
            lines.append(f"Top performers: {g_str}.")
        if losers:
            l_str = ", ".join(f"<b>{c}</b> ({p:+.1f}%)" for c, p in losers)
            lines.append(f"Underperformers: {l_str}.")

        # Spread
        spread = max(returns_by_code.values()) - min(returns_by_code.values())
        if spread > 20:
            lines.append(
                f"Performance spread of <b>{spread:.1f}%</b> between best and worst suggests "
                f"strong stock-picking opportunity within this selection."
            )
        elif spread < 5:
            lines.append(
                f"Tight spread of {spread:.1f}% — companies are moving largely in tandem "
                f"(high correlation)."
            )

    summary_html = " ".join(lines)

    return html.Div([
        html.Div([
            html.Div([
                html.Span(sentiment.upper(), style=dict(
                    background=badge_color + "25", color=badge_color,
                    border=f"1px solid {badge_color}60",
                    borderRadius="20px", padding="2px 12px",
                    fontSize="0.65rem", fontWeight=800,
                    textTransform="uppercase", letterSpacing="0.1em",
                    marginRight="10px",
                )),
                html.Span("Interpretation", style=dict(
                    fontWeight=700, color=C["text"], fontSize="0.88rem",
                )),
            ], style=dict(marginBottom="10px", display="flex", alignItems="center")),
            html.Div(
                dcc.Markdown(
                    summary_html,
                    dangerously_allow_html=True,
                    style=dict(color=C["muted"], fontSize="0.82rem",
                               lineHeight="1.8", margin=0),
                ),
            ),
        ], style=dict(
            background=C["card"], border=f"1px solid {C['border']}",
            borderLeft=f"3px solid {badge_color}",
            borderRadius="10px", padding="18px 24px",
            margin="0 24px 4px",
        )),
    ])


# ── Tab builder ───────────────────────────────────────────────────────────────

def build_explorer_tab():
    from datetime import date, timedelta
    today         = date.today()
    default_start = (today - timedelta(days=30)).isoformat()
    default_end   = today.isoformat()

    code_options = [
        {"label": f"{c}  —  {_NAME_MAP.get(c, '')}", "value": c}
        for c in _ALL_NSE_CODES
    ]

    # Field glossary rows
    glossary_rows = [
        html.Tr([
            html.Td(col, style=dict(fontWeight=700, color=C["accent"], fontSize="0.78rem",
                                    padding="4px 12px 4px 0", whiteSpace="nowrap")),
            html.Td(desc, style=dict(color=C["muted"], fontSize="0.78rem", padding="4px 0")),
        ])
        for col, desc in _FIELD_GLOSSARY.items()
    ]

    guidance = html.Div([
        # Quick-start
        html.Div([
            html.Div("Quick Start", style=dict(fontWeight=700, color=C["text"],
                                               fontSize="0.85rem", marginBottom="8px")),
            html.Ol([
                html.Li("Pick a date range below  (data available: 2000 → today — upload earlier years via template)",
                        style=dict(color=C["muted"], fontSize="0.8rem", marginBottom="4px")),
                html.Li("Select the companies you want  (all 62 are pre-selected)",
                        style=dict(color=C["muted"], fontSize="0.8rem", marginBottom="4px")),
                html.Li("Click  Load Data  — chart, stats and table appear instantly",
                        style=dict(color=C["muted"], fontSize="0.8rem", marginBottom="4px")),
                html.Li("Toggle Candlestick / Line with the chart button",
                        style=dict(color=C["muted"], fontSize="0.8rem", marginBottom="4px")),
                html.Li("Download as CSV or Excel when you're done",
                        style=dict(color=C["muted"], fontSize="0.8rem")),
            ], style=dict(paddingLeft="18px", margin=0)),
        ], style=dict(flex="1", minWidth="280px")),

        # Field glossary
        html.Div([
            html.Div("Column Guide — what every NSE field means",
                     style=dict(fontWeight=700, color=C["text"],
                                fontSize="0.85rem", marginBottom="8px")),
            html.Table(html.Tbody(glossary_rows), style=dict(borderCollapse="collapse")),
        ], style=dict(flex="2", minWidth="340px")),

    ], style=dict(
        display="flex", gap="32px", flexWrap="wrap",
        background=C["panel"], border=f"1px solid {C['border']}",
        borderRadius="10px", padding="18px 24px", margin="16px 24px 0",
    ))

    toggle_btn_style = dict(
        border=f"1px solid {C['border']}", borderRadius="20px",
        padding="6px 14px", cursor="pointer", fontSize="0.8rem",
        fontWeight=600, background=C["card"], color=C["muted"],
    )

    return html.Div([
        # ── Header ──────────────────────────────────────────────────────────
        html.Div([
            html.Div([
                html.Span("📅  NSE Data Explorer",
                          style=dict(fontSize="1.15rem", fontWeight=800, color=C["text"])),
                html.Span(
                    f"  ·  Data available: 2007 – {pd.Timestamp('today').strftime('%d %b %Y')}  "
                    f"·  {len(_ALL_NSE_CODES)} companies",
                    style=dict(fontSize="0.75rem", color=C["muted"], marginLeft="8px"),
                ),
            ]),
            html.Button(
                "▼ Field Guide", id="explorer-guide-toggle", n_clicks=0,
                style={**toggle_btn_style, "background": "transparent"},
            ),
        ], style=dict(
            display="flex", justifyContent="space-between", alignItems="center",
            padding="16px 24px 8px",
            borderBottom=f"1px solid {C['border']}",
        )),

        # ── Collapsible guidance panel ───────────────────────────────────────
        html.Div(guidance, id="explorer-guide-panel", style=dict(display="none")),

        # ── Sticky controls bar ──────────────────────────────────────────────
        html.Div([
            # Date range
            html.Div([
                html.Label("Date Range",
                           style=dict(color=C["muted"], fontSize="0.72rem",
                                      display="block", marginBottom="5px",
                                      textTransform="uppercase", letterSpacing="0.07em")),
                html.Div([
                    html.Span("Quick pick:", style=dict(fontSize="0.7rem", color=C["muted"],
                                                         alignSelf="center", marginRight="4px")),
                    *[html.Button(lbl, id=f"explorer-preset-{key}", n_clicks=0, style=dict(
                        background=C["card"], color=C["accent"],
                        border=f"1px solid {C['border']}", borderRadius="14px",
                        padding="3px 10px", cursor="pointer", fontSize="0.72rem", fontWeight=600,
                        marginRight="4px",
                    )) for lbl, key in [("7D","7d"),("30D","30d"),("3M","3m"),("1Y","1y"),("YTD","ytd"),("Max","max")]
                    ],
                ], style=dict(display="flex", alignItems="center", flexWrap="wrap", marginBottom="6px")),
                html.Div(
                    dcc.DatePickerRange(
                        id="explorer-dates",
                        start_date=default_start,
                        end_date=default_end,
                        min_date_allowed=_ARCHIVE_MIN_DATE,
                        max_date_allowed=_ARCHIVE_MAX_DATE,
                        display_format="DD MMM YYYY",
                        minimum_nights=0,
                        style=dict(fontSize="0.82rem"),
                    ),
                    id="explorer-dates-wrap",
                ),
            ], style=dict(flex="0 0 auto")),

            # Company selector
            html.Div([
                html.Div([
                    html.Label("Companies",
                               style=dict(color=C["muted"], fontSize="0.72rem",
                                          display="block", marginBottom="5px",
                                          textTransform="uppercase", letterSpacing="0.07em")),
                    html.Div([
                        html.Span("All 62 pre-selected",
                                  style=dict(fontSize="0.72rem", color=C["buy"])),
                        html.Button("Select All", id="explorer-select-all", n_clicks=0,
                                    style=dict(marginLeft="10px", background="transparent",
                                               color=C["accent"], border=f"1px solid {C['accent']}",
                                               borderRadius="12px", padding="2px 10px",
                                               cursor="pointer", fontSize="0.72rem")),
                        html.Button("Clear", id="explorer-clear", n_clicks=0,
                                    style=dict(marginLeft="6px", background="transparent",
                                               color=C["muted"], border=f"1px solid {C['border']}",
                                               borderRadius="12px", padding="2px 10px",
                                               cursor="pointer", fontSize="0.72rem")),
                    ], style=dict(display="flex", alignItems="center", marginBottom="5px")),
                ]),
                dcc.Dropdown(
                    id="explorer-codes",
                    options=code_options,
                    value=_ALL_NSE_CODES,
                    multi=True,
                    placeholder="Search and select companies…",
                    style=dict(background=C["card"], fontSize="0.82rem", minWidth="380px"),
                ),
            ], style=dict(flex="1", minWidth="300px")),

            # Action buttons
            html.Div([
                html.Div("Set date range & companies above, then:",
                         style=dict(fontSize="0.68rem", color=C["muted"], marginBottom="5px")),
                html.Button("Load Data", id="explorer-load", n_clicks=0, style=dict(
                    background=C["accent"], color=C["header"],
                    fontWeight=800, border="none", borderRadius="8px",
                    padding="12px 32px", cursor="pointer", fontSize="0.95rem",
                    letterSpacing="0.03em", width="100%",
                )),
            ], style=dict(display="flex", flexDirection="column", alignItems="flex-end", paddingBottom="2px")),

        ], style=dict(
            display="flex", gap="20px", flexWrap="wrap", alignItems="flex-start",
            padding="14px 24px 14px",
            background=C["panel"],
            borderBottom=f"1px solid {C['border']}",
            position="sticky", top="0", zIndex="100",
        )),

        # ── Results area (stats, movers, table) ─────────────────────────────
        dcc.Loading(
            html.Div(
                html.Div(
                    "Select a date range and companies, then click Load Data.",
                    style=dict(color=C["muted"], padding="48px 24px",
                               fontSize="0.9rem", textAlign="center"),
                ),
                id="explorer-results",
            ),
            type="dot", color=C["accent"],
        ),

        # ── Chart section — static, always in DOM ────────────────────────────
        html.Div([
            html.Div([
                html.Div("Price Chart",
                         style=dict(fontWeight=700, color=C["text"], fontSize="0.9rem")),
                html.Div([
                    html.Div("Chart type:", style=dict(
                        color=C["muted"], fontSize="0.78rem",
                        marginRight="8px", alignSelf="center",
                    )),
                    html.Button("Line", id="explorer-toggle-line", n_clicks=0, style=dict(
                        background=C["accent"], color=C["header"],
                        border=f"1px solid {C['border']}", borderRadius="6px 0 0 6px",
                        padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
                    )),
                    html.Button("Candlestick", id="explorer-toggle-candle", n_clicks=0, style=dict(
                        background=C["card"], color=C["muted"],
                        border=f"1px solid {C['border']}", borderLeft="none",
                        borderRadius="0 6px 6px 0",
                        padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
                    )),
                ], style=dict(display="flex", alignItems="center")),
            ], style=dict(display="flex", justifyContent="space-between",
                          alignItems="center", marginBottom="10px")),
            dcc.Graph(
                id="explorer-chart",
                figure=go.Figure(layout=dict(
                    paper_bgcolor=C["card"], plot_bgcolor=C["card"], height=480,
                    xaxis=dict(visible=False), yaxis=dict(visible=False),
                    annotations=[dict(text="Load data to see the chart", showarrow=False,
                                     font=dict(color=C["muted"], size=14))],
                )),
                config=dict(displayModeBar=True, scrollZoom=True,
                            modeBarButtonsToRemove=["lasso2d", "select2d"]),
            ),
        ], id="explorer-chart-section",
           style=dict(display="none", padding="16px 24px",
                      borderBottom=f"1px solid {C['border']}")),

        # ── Export bar — static, always in DOM ───────────────────────────────
        html.Div([
            html.Button("Download CSV", id="explorer-dl-csv", n_clicks=0, style=dict(
                background="transparent", color=C["buy"],
                border=f"1px solid {C['buy']}", borderRadius="8px",
                padding="8px 20px", cursor="pointer", fontSize="0.83rem", fontWeight=600,
            )),
            html.Button("Download Excel (.xlsx)", id="explorer-dl-excel", n_clicks=0, style=dict(
                background="transparent", color=C["accent"],
                border=f"1px solid {C['accent']}", borderRadius="8px",
                padding="8px 20px", cursor="pointer", fontSize="0.83rem", fontWeight=600,
            )),
        ], id="explorer-export-bar",
           style=dict(display="none", gap="12px", padding="16px 24px 32px")),

    ], style=dict(overflowY="auto", maxHeight="calc(100vh - 105px)"))


# ── Results renderer ─────────────────────────────────────────────────────────

def _render_results(df):
    if df.empty:
        return html.Div(
            "No data found for this selection. Try a wider date range or different companies.",
            style=dict(color=C["muted"], padding="48px 24px", textAlign="center"),
        )

    stats, movers = _compute_stats(df)
    n_codes = df["Code"].nunique()
    min_dt  = df["_dt"].min().strftime("%d %b %Y")
    max_dt  = df["_dt"].max().strftime("%d %b %Y")

    # ── Summary stats strip ──────────────────────────────────────────────────
    if n_codes == 1:
        # Single company: show period return rather than best/worst comparison
        code = df["Code"].iloc[0]
        prices = _to_num(df.sort_values("_dt")["Day Price"]).dropna()
        period_ret = (prices.iloc[-1] / prices.iloc[0] - 1) * 100 if len(prices) >= 2 else 0
        ret_color = C["buy"] if period_ret >= 0 else C["sell"]
        stat_cards = [
            _stat_card("Period Return",
                       f"{period_ret:+.1f}%",
                       ret_color,
                       f"{_NAME_MAP.get(code, code)}  ·  {min_dt} → {max_dt}"),
            _stat_card("Biggest Single-Day Move",
                       f"{stats['biggest_day_pct']:+.1f}%",
                       C["hold"],
                       stats["biggest_day_info"]),
            _stat_card("Total Records",
                       f"{stats['total_records']:,}",
                       C["accent"],
                       f"1 company  ·  {min_dt} → {max_dt}"),
        ]
    else:
        stat_cards = [
            _stat_card("Best Performer",
                       f"{stats['best_code']}  {stats['best_pct']:+.1f}%",
                       C["buy"],
                       _NAME_MAP.get(stats["best_code"], "")),
            _stat_card("Worst Performer",
                       f"{stats['worst_code']}  {stats['worst_pct']:+.1f}%",
                       C["sell"],
                       _NAME_MAP.get(stats["worst_code"], "")),
            _stat_card("Biggest Single-Day Move",
                       f"{stats['biggest_day_pct']:+.1f}%",
                       C["hold"],
                       stats["biggest_day_info"]),
            _stat_card("Total Records",
                       f"{stats['total_records']:,}",
                       C["accent"],
                       f"{n_codes} companies  ·  {min_dt} → {max_dt}"),
        ]

    stats_strip = html.Div(stat_cards, style=dict(
        display="flex", gap="12px", flexWrap="wrap",
        padding="16px 24px", borderBottom=f"1px solid {C['border']}",
    ))

    # ── Top Movers strip (hidden for single-company) ─────────────────────────
    mover_badges = []
    gainers = [(c, p) for t, c, p in movers if t == "gain"]
    losers  = [(c, p) for t, c, p in movers if t == "loss"]
    if gainers or losers:
        parts = []
        if gainers:
            parts += [
                html.Div("Top Gainers", style=dict(
                    fontSize="0.68rem", color=C["buy"], fontWeight=700,
                    textTransform="uppercase", letterSpacing="0.08em",
                    marginRight="4px", alignSelf="center",
                )),
                *[_mover_badge(c, p, True) for c, p in gainers],
            ]
        if gainers and losers:
            parts.append(html.Div("", style=dict(
                width="1px", background=C["border"],
                margin="0 12px", alignSelf="stretch",
            )))
        if losers:
            parts += [
                html.Div("Top Losers", style=dict(
                    fontSize="0.68rem", color=C["sell"], fontWeight=700,
                    textTransform="uppercase", letterSpacing="0.08em",
                    marginRight="4px", alignSelf="center",
                )),
                *[_mover_badge(c, p, False) for c, p in losers],
            ]
        mover_badges = parts

    movers_strip = html.Div(mover_badges, style=dict(
        display="flex", gap="8px", flexWrap="wrap", alignItems="center",
        padding="12px 24px", borderBottom=f"1px solid {C['border']}",
        background=C["panel"],
    )) if mover_badges else html.Div()

    # ── Interpretation panel ─────────────────────────────────────────────────
    interpretation = _build_interpretation(df, stats, movers)

    # ── Data table — all 13 NSE columns ─────────────────────────────────────
    df_disp = df.copy()
    df_disp["Date"] = df_disp["_dt"].dt.strftime("%d/%m/%Y")
    cols_present = [c for c in _NSE_COLS_ORDERED if c in df_disp.columns]
    df_disp = df_disp[cols_present]

    table_section = html.Div([
        html.Div([
            html.Div("Full NSE Data  —  all 13 columns",
                     style=dict(fontWeight=700, color=C["text"], fontSize="0.88rem")),
            html.Div("Sort any column · Filter using the search row · 100 rows per page",
                     style=dict(fontSize="0.72rem", color=C["muted"], marginTop="2px")),
        ], style=dict(marginBottom="12px")),

        dash_table.DataTable(
            id="explorer-table",
            data=df_disp.to_dict("records"),
            columns=[{"name": c, "id": c} for c in cols_present],
            page_size=100,
            sort_action="native",
            filter_action="native",
            style_table=dict(overflowX="auto", minWidth="100%"),
            style_header=dict(
                background=C["panel"], color=C["muted"],
                fontWeight=700, fontSize="0.7rem",
                border=f"1px solid {C['border']}",
                textTransform="uppercase", letterSpacing="0.05em",
                whiteSpace="nowrap",
            ),
            style_cell=dict(
                background=C["card"], color=C["text"],
                fontSize="0.82rem", padding="7px 12px",
                border=f"1px solid {C['border']}",
                fontVariantNumeric="tabular-nums",
                whiteSpace="nowrap",
            ),
            style_data_conditional=[
                {"if": {"column_id": "Day Price"}, "fontWeight": 700, "color": C["accent"]},
                {"if": {"column_id": "Code"},      "fontWeight": 700},
                {"if": {"column_id": "Change",  "filter_query": "{Change} > 0"},
                 "color": C["buy"],  "fontWeight": 600},
                {"if": {"column_id": "Change",  "filter_query": "{Change} < 0"},
                 "color": C["sell"], "fontWeight": 600},
                {"if": {"column_id": "Change%", "filter_query": "{Change%} > 0"},
                 "color": C["buy"],  "fontWeight": 600},
                {"if": {"column_id": "Change%", "filter_query": "{Change%} < 0"},
                 "color": C["sell"], "fontWeight": 600},
            ],
        ),

    ], style=dict(padding="16px 24px 32px"))

    return html.Div([stats_strip, movers_strip, interpretation, table_section])


# ── Explorer Callbacks ────────────────────────────────────────────────────────

@app.callback(
    Output("explorer-guide-panel", "style"),
    Input("explorer-guide-toggle", "n_clicks"),
    prevent_initial_call=True,
)
def toggle_guide(n):
    return dict(display="block") if n % 2 == 1 else dict(display="none")


@app.callback(
    Output("explorer-dates", "start_date"),
    Output("explorer-dates", "end_date"),
    Input("explorer-preset-7d",  "n_clicks"),
    Input("explorer-preset-30d", "n_clicks"),
    Input("explorer-preset-3m",  "n_clicks"),
    Input("explorer-preset-1y",  "n_clicks"),
    Input("explorer-preset-ytd", "n_clicks"),
    Input("explorer-preset-max", "n_clicks"),
    prevent_initial_call=True,
)
def explorer_date_preset(_7d, _30d, _3m, _1y, _ytd, _max):
    from datetime import date, timedelta
    today = date.today()
    presets = {
        "explorer-preset-7d":  (today - timedelta(days=7),   today),
        "explorer-preset-30d": (today - timedelta(days=30),  today),
        "explorer-preset-3m":  (today - timedelta(days=91),  today),
        "explorer-preset-1y":  (today - timedelta(days=365), today),
        "explorer-preset-ytd": (date(today.year, 1, 1),      today),
        "explorer-preset-max": (date(2000, 1, 1),            today),
    }
    start, end = presets.get(ctx.triggered_id, (today - timedelta(days=30), today))
    return start.isoformat(), end.isoformat()


@app.callback(
    Output("explorer-codes", "value"),
    Input("explorer-select-all", "n_clicks"),
    Input("explorer-clear",      "n_clicks"),
    prevent_initial_call=True,
)
def explorer_code_buttons(n_all, n_clear):
    return _ALL_NSE_CODES if ctx.triggered_id == "explorer-select-all" else []


_CHART_VISIBLE  = dict(display="block", padding="16px 24px", borderBottom=f"1px solid {C['border']}")
_CHART_HIDDEN   = dict(display="none",  padding="16px 24px", borderBottom=f"1px solid {C['border']}")
_EXPORT_VISIBLE = dict(display="flex",  gap="12px", padding="16px 24px 32px")
_EXPORT_HIDDEN  = dict(display="none",  gap="12px", padding="16px 24px 32px")

_LINE_ACTIVE = dict(
    background=C["accent"], color=C["header"],
    border=f"1px solid {C['border']}", borderRadius="6px 0 0 6px",
    padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
)
_LINE_INACTIVE = dict(
    background=C["card"], color=C["muted"],
    border=f"1px solid {C['border']}", borderRadius="6px 0 0 6px",
    padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
)
_CANDLE_ACTIVE = dict(
    background=C["accent"], color=C["header"],
    border=f"1px solid {C['border']}", borderLeft="none", borderRadius="0 6px 6px 0",
    padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
)
_CANDLE_INACTIVE = dict(
    background=C["card"], color=C["muted"],
    border=f"1px solid {C['border']}", borderLeft="none", borderRadius="0 6px 6px 0",
    padding="5px 14px", cursor="pointer", fontSize="0.8rem", fontWeight=600,
)


@app.callback(
    Output("explorer-results",       "children"),
    Output("explorer-data",          "data"),
    Output("explorer-chart-section", "style"),
    Output("explorer-export-bar",    "style"),
    Input("explorer-load", "n_clicks"),
    State("explorer-dates", "start_date"),
    State("explorer-dates", "end_date"),
    State("explorer-codes", "value"),
    prevent_initial_call=True,
)
def explorer_load(n_load, start, end, codes):
    if not start or not end:
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update

    selected = codes or _ALL_NSE_CODES
    try:
        df = _load_archive_range(selected, start, end)
    except Exception as e:
        return (
            html.Div(f"Error loading data: {e}",
                     style=dict(color=C["sell"], padding="24px")),
            {}, _CHART_HIDDEN, _EXPORT_HIDDEN,
        )

    if df.empty:
        return (
            html.Div("No records found for this selection.",
                     style=dict(color=C["muted"], padding="48px 24px", textAlign="center")),
            {}, _CHART_HIDDEN, _EXPORT_HIDDEN,
        )

    records = df.assign(_dt=df["_dt"].astype(str)).to_dict("records")
    return _render_results(df), records, _CHART_VISIBLE, _EXPORT_VISIBLE


@app.callback(
    Output("explorer-chart-type",    "data"),
    Output("explorer-toggle-line",   "style"),
    Output("explorer-toggle-candle", "style"),
    Input("explorer-toggle-line",   "n_clicks"),
    Input("explorer-toggle-candle", "n_clicks"),
    prevent_initial_call=True,
)
def explorer_chart_toggle(_n_line, _n_candle):
    if ctx.triggered_id == "explorer-toggle-candle":
        return "candlestick", _LINE_INACTIVE, _CANDLE_ACTIVE
    return "line", _LINE_ACTIVE, _CANDLE_INACTIVE


@app.callback(
    Output("explorer-chart", "figure"),
    Input("explorer-data",       "data"),
    Input("explorer-chart-type", "data"),
    prevent_initial_call=True,
)
def explorer_draw_chart(records, chart_type):
    if not records:
        return go.Figure()
    df = pd.DataFrame(records)
    df["_dt"] = pd.to_datetime(df["_dt"])
    return _build_chart(df, chart_type or "line")


@app.callback(
    Output("explorer-download-csv", "data"),
    Input("explorer-dl-csv", "n_clicks"),
    State("explorer-data", "data"),
    prevent_initial_call=True,
)
def explorer_dl_csv(n, records):
    if not n or not records:
        return dash.no_update
    df = pd.DataFrame(records)
    cols = [c for c in _NSE_COLS_ORDERED if c in df.columns]
    return dcc.send_data_frame(df[cols].to_csv, "nse_data.csv", index=False)


@app.callback(
    Output("explorer-download-excel", "data"),
    Input("explorer-dl-excel", "n_clicks"),
    State("explorer-data", "data"),
    prevent_initial_call=True,
)
def explorer_dl_excel(n, records):
    if not n or not records:
        return dash.no_update
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    df = pd.DataFrame(records)
    cols = [c for c in _NSE_COLS_ORDERED if c in df.columns]
    df = df[cols]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NSE Data", index=False)
        ws = writer.sheets["NSE Data"]

        # Style header row
        header_fill = PatternFill("solid", fgColor="0D1117")
        accent_font = Font(bold=True, color="38BDF8")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = accent_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf.seek(0)
    return dcc.send_bytes(buf.read, "nse_data.xlsx")


# ── run_quick_fn ──────────────────────────────────────────────────────────────
def run_quick_fn(df_clean, ticker):
    from src.analysis.returns import daily_return_analysis
    from src.analysis.moving_averages import compute_moving_averages
    from src.analysis.risk import value_at_risk
    from src.features.engineer import build_feature_matrix, select_top_features
    from src.models.xgboost_model import train_xgboost, save_xgboost
    from src.models.ensemble import generate_signal
    from sklearn.metrics import mean_squared_error, mean_absolute_error
    ret_df, _ = daily_return_analysis(df_clean)
    ma_df = compute_moving_averages(ret_df)
    var_r = value_at_risk(df_clean, investment=DEFAULT_INVESTMENT, confidence=DEFAULT_CONFIDENCE)
    feat  = build_feature_matrix(ma_df)
    fcols = select_top_features(feat)
    model, _, actuals, preds = train_xgboost(feat, fcols)
    save_xgboost(model, ticker)
    curr  = float(df_clean["Close"].iloc[-1])
    var_p = var_r["historical_var_pct"]
    sig   = generate_signal(curr, float(preds[-1]), var_p)
    rmse  = float(np.sqrt(mean_squared_error(actuals, preds)))
    mape  = float(np.mean(np.abs((actuals-preds)/np.where(actuals==0,1,actuals)))*100)
    dacc  = float(np.mean(np.sign(np.diff(actuals))==np.sign(np.diff(preds)))*100)
    result = {**sig, "metrics": dict(rmse=rmse, mape=mape, directional_accuracy=dacc),
              "actuals": actuals.tolist(), "preds": preds.tolist(), "ma_df": ma_df}
    from main import _save_signal
    _save_signal(ticker, {k:v for k,v in result.items() if k != "ma_df"})
    return result


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="NSE Market Intelligence Platform")
    parser.add_argument("--host", default="127.0.0.1",
                        help="Host to bind (use 0.0.0.0 for network access)")
    parser.add_argument("--port", type=int, default=8050)
    args = parser.parse_args()
    app.run(debug=False, host=args.host, port=args.port)


